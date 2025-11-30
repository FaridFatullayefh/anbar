from flask import Flask, render_template, request, redirect, url_for, session, flash,send_file,make_response,Response
import sqlite3
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import datetime
import pytz
import json
import openpyxl
import os
import time
import logging
import random
import string
from io import BytesIO 
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
logging.basicConfig(level=logging.INFO)
app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET', 'secret_key_lokal')
import sys
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

def get_db_path():

    DB_FILENAME = 'warehouse.db' 
    
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), DB_FILENAME) 
    
    return os.path.join(os.path.dirname(__file__), DB_FILENAME)

db_path = get_db_path()

AZERBAIJAN_TIMEZONE = pytz.timezone('Asia/Baku')

class DatabaseConnection:
    def __init__(self):
        self.conn = None

    def __enter__(self):
        self.conn = sqlite3.connect(
            db_path,
            detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES,
            timeout=30.0,
            check_same_thread=False
        )
        self.conn.row_factory = sqlite3.Row
        self.conn.execute('PRAGMA journal_mode=WAL')
        self.conn.execute('PRAGMA foreign_keys=ON') 
        self.conn.execute('PRAGMA busy_timeout=30000')
        return self.conn

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn:
            self.conn.close()

def get_db_connection():
    conn = sqlite3.connect(
        db_path,
        detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES,
        timeout=30.0,
        isolation_level=None,
        check_same_thread=False
    )
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA busy_timeout=30000')
    conn.execute('PRAGMA foreign_keys=ON')
    return conn

def get_current_baku_time():
    utc_now = datetime.datetime.now(datetime.timezone.utc)
    baku_now = utc_now.astimezone(AZERBAIJAN_TIMEZONE)
    return baku_now.strftime('%Y-%m-%d %H:%M:%S')

def calculate_company_debt(conn, company_id):
    total_debt = conn.execute(
        'SELECT COALESCE(SUM(amount), 0) FROM company_debt_transactions WHERE company_id = ?', 
        (company_id,)
    ).fetchone()[0]
    
    total_payments = conn.execute(
        'SELECT COALESCE(SUM(amount), 0) FROM company_debt_payments WHERE company_id = ?',
        (company_id,)
    ).fetchone()[0]
    
    return total_debt - total_payments


def update_company_debt(conn, company_id):
    total_debt = calculate_company_debt(conn, company_id)
    current_time = get_current_baku_time()
    
    conn.execute(
        'UPDATE company_total_debts SET total_debt = ?, updated_at = ? WHERE id = ?',
        (total_debt, current_time, company_id)
    )
    conn.commit()
    return total_debt

def audit_log(conn, employee_id, action, details=None):
    """
    Enhanced audit logging with detailed information
    
    Args:
        conn: Database connection
        employee_id: ID of the employee performing the action
        action: Main action description
        details: Dictionary containing additional details (will be stored as JSON)
    """
    timestamp = get_current_baku_time()
    details_json = json.dumps(details, ensure_ascii=False) if details else None
    
    conn.execute(
        """
        INSERT INTO audit_log 
        (employee_id, action, details, timestamp) 
        VALUES (?, ?, ?, ?)
        """,
        (employee_id, action, details_json, timestamp)
    )

def record_transaction(conn, product_id, type_, quantity, price, total, employee_id, note='', customer_name='', supplier_id=None):
    """
    Anbar hərəkətlərini 'transactions' cədvəlinə qeyd edir.
    Alış zamanı supplier_id, Satış zamanı customer_name istifadə olunur.
    """
    timestamp = get_current_baku_time()
    
    # 1. SQL Sorğusuna yeni sahəni (supplier_id) əlavə edirik
    conn.execute(
        '''INSERT INTO transactions 
           (product_id, type, quantity, price, total, employee_id, timestamp, note, customer_name, supplier_id) 
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
        (product_id, type_, quantity, price, total, employee_id, timestamp, note, customer_name, supplier_id) 
        # 2. Yeni parametr dəyərini (supplier_id) göndəririk
    )
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Giriş etməlisiniz.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('role') not in [role, 'Admin']: 
                flash('Bu səhifəyə giriş icazəniz yoxdur.', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator
def init_db():
    """Verilənlər bazasının strukturunu yaradır və ya mövcud bazaya lazım olan
    sütunları və default istifadəçini əlavə edir (migrasiya)."""
    
    # generate_password_hash funksiyası mövcud olduğu güman edilir
    global generate_password_hash
    # Əgər generate_password_hash import edilməyibsə:
    # try:
    #     from werkzeug.security import generate_password_hash
    # except ImportError:
    #     def generate_password_hash(password):
    #         return password # Sadəcə mətn şəklində saxlayır (təhlükəsiz deyil, amma işlədir)

    try:
        # Tutaq ki, DatabaseConnection context manager-i mövcuddur.
        with DatabaseConnection() as conn: 
            cursor = conn.cursor()
            
            cursor.execute("PRAGMA foreign_keys = OFF") 

            # 0. AUDIT_LOGS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS audit_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER,
                    action TEXT NOT NULL,
                    details TEXT,
                    ip_address TEXT,
                    user_agent TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
                )
            ''')
            
            # 1. EMPLOYEES
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS employees (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    role TEXT NOT NULL
                )
            ''')
            
            # 2. SUPPLIERS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS suppliers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    contact_person TEXT,
                    phone TEXT,
                    email TEXT,
                    address TEXT,
                    tax_id TEXT,
                    bank_account TEXT,
                    notes TEXT,
                    is_active INTEGER DEFAULT 1,
                    created_at TEXT DEFAULT (datetime('now')),
                    updated_at TEXT
                )
            ''')
            
            # 3. PRODUCTS
            conn.execute("""
                CREATE TABLE IF NOT EXISTS products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    sku TEXT UNIQUE NOT NULL,
                    category TEXT,
                    unit TEXT,
                    stock REAL NOT NULL DEFAULT 0,
                    min_stock REAL DEFAULT 0,
                    price REAL DEFAULT 0,
                    supplier_id INTEGER,
                    location TEXT,
                    description TEXT,
                    is_active INTEGER DEFAULT 1,
                    created_at TEXT DEFAULT (datetime('now')),
                    updated_at TEXT,
                    FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
                )
            """)

            # 4. COMPANY_TOTAL_DEBTS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS company_total_debts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    company_name TEXT UNIQUE NOT NULL,
                    total_debt REAL NOT NULL DEFAULT 0,
                    last_updated TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    updated_at TEXT
                )
            ''')

            # 5. COMPANY_DEBT_TRANSACTIONS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS company_debt_transactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    company_id INTEGER,
                    amount REAL NOT NULL, 
                    description TEXT,
                    transaction_date TEXT,
                    employee_id INTEGER,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (company_id) REFERENCES company_total_debts(id) ON DELETE CASCADE,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
                )
            ''')

            # 6. COMPANY_DEBT_PAYMENTS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS company_debt_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    company_id INTEGER,
                    amount REAL NOT NULL, 
                    payment_date TEXT,
                    description TEXT,
                    employee_id INTEGER,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (company_id) REFERENCES company_total_debts(id) ON DELETE CASCADE,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
                )
            ''')
            
            # 7. COMPANY_PAYMENTS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS company_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    company_id INTEGER NOT NULL,
                    amount REAL NOT NULL,
                    payment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    description TEXT,
                    employee_id INTEGER NOT NULL,
                    FOREIGN KEY (company_id) REFERENCES company_total_debts(id),
                    FOREIGN KEY (employee_id) REFERENCES employees(id)
                );
            ''')
            
            # 8. PRODUCTION_TRANSFERS
            conn.execute("""
                CREATE TABLE IF NOT EXISTS production_transfers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    transfer_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    user_id INTEGER,
                    description TEXT,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES employees(id) ON DELETE SET NULL
                )
            """)

            # 9. PRODUCTION_TRANSFER_ITEMS
            conn.execute("""
                CREATE TABLE IF NOT EXISTS production_transfer_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    transfer_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    quantity REAL NOT NULL,
                    type TEXT NOT NULL CHECK (type IN ('Input', 'Output')),
                    FOREIGN KEY (transfer_id) REFERENCES production_transfers(id) ON DELETE CASCADE,
                    FOREIGN KEY (product_id) REFERENCES products(id)
                )
            """)

            # 10. AUDIT_LOG
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS audit_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER,
                    action TEXT,
                    details TEXT,
                    timestamp TEXT,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
                )
            ''')
            
            # Add details column if it doesn't exist
            cursor.execute("PRAGMA table_info(audit_log)")
            columns = [column[1] for column in cursor.fetchall()]
            if 'details' not in columns:
                cursor.execute('ALTER TABLE audit_log ADD COLUMN details TEXT')
            
            # 11. TRANSACTIONS - CUSTOMER_NAME və NOTE SÜTUNLARI ƏLAVƏ EDİLDİ
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS transactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER,
                    type TEXT,
                    quantity REAL, 
                    price REAL,
                    total REAL,
                    employee_id INTEGER,
                    timestamp TEXT,
                    customer_name TEXT,
                    note TEXT, 
                    supplier_id INTEGER,
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE SET NULL,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL,
                    FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
                )
            ''')
            
            # TRANSACTIONS CƏDVƏLİ ÜÇÜN MİQRASİYA
            
            # 1. 'note' sütununu yoxla və əlavə et
            try:
                cursor.execute("SELECT note FROM transactions LIMIT 1")
            except sqlite3.OperationalError:
                print(">>> 'transactions' cədvəlində 'note' sütunu yoxdur. Əlavə edilir...")
                conn.execute("ALTER TABLE transactions ADD COLUMN note TEXT")
                conn.commit()
                print("✓ 'note' sütunu əlavə edildi.")

            # 2. 'customer_name' sütununu yoxla və əlavə et (Xətanı düzəldir)
            try:
                cursor.execute("SELECT customer_name FROM transactions LIMIT 1")
            except sqlite3.OperationalError:
                print(">>> 'transactions' cədvəlində 'customer_name' sütunu yoxdur. Əlavə edilir...")
                conn.execute("ALTER TABLE transactions ADD COLUMN customer_name TEXT")
                conn.commit()
                print("✓ 'customer_name' sütunu əlavə edildi.")
            
            # 3. 'supplier_id' sütununu yoxla və əlavə et
            try:
                cursor.execute("SELECT supplier_id FROM transactions LIMIT 1")
            except sqlite3.OperationalError:
                print(">>> 'transactions' cədvəlində 'supplier_id' sütunu yoxdur. Əlavə edilir...")
                conn.execute("ALTER TABLE transactions ADD COLUMN supplier_id INTEGER REFERENCES suppliers(id) ON DELETE SET NULL")
                conn.commit()
                print("✓ 'supplier_id' sütunu əlavə edildi.")
            
            # 12. FINANCIAL_TRANSACTIONS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS financial_transactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    type TEXT NOT NULL,
                    category TEXT,
                    amount REAL NOT NULL,
                    description TEXT,
                    timestamp TEXT,
                    employee_id INTEGER,
                    supplier_id INTEGER,
                    customer_name TEXT,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL,
                    FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
                )
            ''')
            
            # Check if customer_name column exists in financial_transactions
            cursor.execute("PRAGMA table_info(financial_transactions)")
            columns = [column[1] for column in cursor.fetchall()]
            if 'customer_name' not in columns:
                print(">>> 'financial_transactions' cədvəlində 'customer_name' sütunu yoxdur. Əlavə edilir...")
                cursor.execute("ALTER TABLE financial_transactions ADD COLUMN customer_name TEXT")
                conn.commit()
                print("✓ 'customer_name' sütunu əlavə edildi.")
            
            # 13. DEBTS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS debts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    party_name TEXT,
                    type TEXT,
                    initial_amount REAL,
                    paid_amount REAL DEFAULT 0,
                    description TEXT,
                    status TEXT,
                    due_date TEXT,
                    created_at TEXT,
                    employee_id INTEGER,
                    supplier_id INTEGER,
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL,
                    FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
                )
            ''')
            
            # Check if supplier_id column exists in debts
            cursor.execute("PRAGMA table_info(debts)")
            columns = [column[1] for column in cursor.fetchall()]
            if 'supplier_id' not in columns:
                print(">>> 'debts' cədvəlində 'supplier_id' sütunu yoxdur. Əlavə edilir...")
                cursor.execute("ALTER TABLE debts ADD COLUMN supplier_id INTEGER REFERENCES suppliers(id) ON DELETE SET NULL")
                conn.commit()
                print("✓ 'supplier_id' sütunu əlavə edildi.")

            # 14. CUSTOMER_DEBTS
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS customer_debts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_name TEXT,
                    amount REAL,
                    description TEXT,
                    status TEXT,
                    created_by INTEGER,
                    created_at TEXT,
                    paid_amount REAL DEFAULT 0.0, 
                    FOREIGN KEY (created_by) REFERENCES employees(id) ON DELETE SET NULL
                )
            ''')
            
            # 15. DEBT_PAYMENTS (DÜZƏLİŞ)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS debt_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    debt_id INTEGER, 
                    amount REAL NOT NULL,
                    payment_date TEXT,
                    description TEXT,
                    employee_id INTEGER,
                    created_at TEXT DEFAULT (datetime('now')),
                    source_table TEXT NOT NULL DEFAULT 'debts', 
                    transaction_type TEXT,  -- <<< ƏSAS DÜZƏLİŞ: transaction_type əlavə edildi
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
                )
            ''')
            
            # DEBT_PAYMENTS CƏDVƏLİ ÜÇÜN MİQRASİYA (Köhnə DB üçün)
            try:
                cursor.execute("SELECT transaction_type FROM debt_payments LIMIT 1")
            except sqlite3.OperationalError:
                print(">>> 'debt_payments' cədvəlində 'transaction_type' sütunu yoxdur. Əlavə edilir...")
                conn.execute("ALTER TABLE debt_payments ADD COLUMN transaction_type TEXT")
                conn.commit()
                print("✓ 'transaction_type' sütunu əlavə edildi.")
            
            
            # 16. PRODUCT_STOCK
            conn.execute("""
                CREATE TABLE IF NOT EXISTS product_stock (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    quantity REAL NOT NULL,
                    stock_type TEXT,
                    transaction_id INTEGER,
                    transaction_type TEXT,
                    transaction_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(product_id) REFERENCES products(id)
                )
            """)
            
            # KÖHNƏ 'users' istinadını 'employees' olaraq yeniləyən köçürmə məntiqi (dəyişilmədi)
            try:
                # ... (production_transfers update logic) ...
                sql_result = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='production_transfers'").fetchone()
                
                if sql_result:
                    sql = sql_result[0]
                    if "references users" in sql.lower():
                        
                        print(">>> production_transfers cədvəlində köhnə 'users' istinadı aşkarlandı. Düzəldilir...")

                        conn.execute("""
                            CREATE TABLE production_transfers_temp AS 
                            SELECT id, transfer_date, user_id, description, created_at FROM production_transfers
                        """)
                        
                        conn.execute("DROP TABLE production_transfers")

                        conn.execute("""
                            CREATE TABLE production_transfers (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                transfer_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                                user_id INTEGER,
                                description TEXT,
                                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY (user_id) REFERENCES employees(id) ON DELETE SET NULL
                            )
                        """)
                        
                        conn.execute("""
                            INSERT INTO production_transfers (id, transfer_date, user_id, description, created_at)
                            SELECT id, transfer_date, user_id, description, created_at FROM production_transfers_temp
                        """)
                        
                        conn.execute("DROP TABLE production_transfers_temp")
                        print("✓ production_transfers cədvəlinin FOREIGN KEY istinadı EMPLOYEES olaraq yeniləndi.")

            except Exception as e:
                # print(f"production_transfers yeniləməsi zamanı kritik xəta: {e}") 
                pass 

            cursor.execute("PRAGMA foreign_keys = ON") 

            # Admin istifadəçisini yoxlayıb əlavə etmək (ŞİFRƏ 'admin' OLARAQ QALIR)
            cursor.execute('SELECT * FROM employees WHERE username = ?', ('admin',))
            if not cursor.fetchone():
                # 'os' və 'generate_password_hash' mövcud olduğu güman edilir
                default_admin_pw = os.environ.get('ADMIN_PASSWORD', 'admin')
                hashed_pw = generate_password_hash(default_admin_pw) 
                cursor.execute('INSERT INTO employees (username, password, role) VALUES (?, ?, ?)',
                               ('admin', hashed_pw, 'Admin'))
                print(f'✓ Admin user created: username=admin, password={default_admin_pw}')

            conn.commit()
            
    except Exception as e:
        print(f"Verilənlər bazası başlatma zamanı ümumi xəta: {e}")
def transliterate_az(text):
    if not text:
        return ""
    
    transliteration_map = {
        'ə': 'e', 'Ə': 'E',
        'ç': 'c', 'Ç': 'C',
        'ğ': 'g', 'Ğ': 'G',
        'ı': 'i', 'I': 'I', 
        'ö': 'o', 'Ö': 'O',
        'ş': 's', 'Ş': 'S',
        'ü': 'u', 'Ü': 'U',
    }
    
    for original, replacement in transliteration_map.items():
        text = text.replace(original, replacement)
    
    return text
@app.route('/')
@login_required
def index():
    conn = get_db_connection()
    try:
        income_row = conn.execute(
            "SELECT SUM(amount) AS total FROM financial_transactions WHERE type = 'Income'"
        ).fetchone()
        expense_row = conn.execute(
            "SELECT SUM(amount) AS total FROM financial_transactions WHERE type = 'Expense'"
        ).fetchone()
        total_income = income_row['total'] or 0
        total_expense = expense_row['total'] or 0
        net_balance = total_income - total_expense

        debts_row = conn.execute("""
            SELECT 
                SUM(CASE WHEN type = 'Debt' THEN (initial_amount - paid_amount) ELSE 0 END) AS total_debts,
                SUM(CASE WHEN type = 'Receivable' THEN (initial_amount - paid_amount) ELSE 0 END) AS total_receivables
            FROM debts
            WHERE (initial_amount - paid_amount) > 0.01
        """).fetchone()
        total_debts = debts_row['total_debts'] or 0
        total_receivables_from_debts = debts_row['total_receivables'] or 0

        customer_debts_row = conn.execute("""
            SELECT SUM(amount - paid_amount) AS total_customer_debts
            FROM customer_debts
            WHERE (amount - paid_amount) > 0.01
        """).fetchone()
        total_receivables_from_customers = customer_debts_row['total_customer_debts'] or 0
        total_receivables = total_receivables_from_debts + total_receivables_from_customers

        latest_transactions = conn.execute("""
            SELECT id, type, category, amount, description, timestamp, employee_id 
            FROM financial_transactions 
            ORDER BY timestamp DESC 
            LIMIT 5
        """).fetchall()

        processed_transactions = []
        for t in latest_transactions:
            t_dict = dict(t)
            try:
                t_dict['timestamp'] = datetime.datetime.strptime(t['timestamp'], '%Y-%m-%d %H:%M:%S')
            except ValueError:
                pass 
            processed_transactions.append(t_dict)

    finally:
        conn.close()

    return render_template(
        'index.html',
        username=session.get('username'),
        net_balance=net_balance,
        total_debts=total_debts,
        total_receivables=total_receivables,
        latest_transactions=processed_transactions 
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = get_db_connection()
        try:
            user = conn.execute('SELECT * FROM employees WHERE username = ?', (username,)).fetchone()
        finally:
            conn.close()
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash('Uğurla daxil oldunuz!', 'success')
            return redirect(url_for('index'))
        else:
            flash('İstifadəçi adı və ya şifrə səhvdir.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Çıxış edildi.', 'info')
    return redirect(url_for('login'))


@app.route('/employees', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def employees():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role') or 'Operator'
        conn = get_db_connection()
        try:
            hashed = generate_password_hash(password)
            conn.execute('INSERT INTO employees (username, password, role) VALUES (?, ?, ?)', (username, hashed, role))
            audit_log(conn, session['user_id'], f'Yeni işçi əlavə edildi: {username}')
            conn.commit()
            flash('İşçi əlavə edildi.', 'success')
        except sqlite3.IntegrityError:
            flash('İstifadəçi adı artıq mövcuddur.', 'danger')
        finally:
            conn.close()

    conn = get_db_connection()
    try:
        employees_list = conn.execute('SELECT id, username, role FROM employees ORDER BY id').fetchall()
    finally:
        conn.close()
    return render_template('employees.html', employees=employees_list)

@app.route('/delete_employee/<int:id>', methods=['POST', 'GET'])
@login_required
@role_required('Admin')
def delete_employee(id):
    conn = get_db_connection()
    try:
        emp = conn.execute('SELECT * FROM employees WHERE id = ?', (id,)).fetchone()
        if not emp:
            flash('İşçi tapılmadı.', 'danger')
            return redirect(url_for('employees'))
        if emp['username'] == 'admin':
            flash('Super admin silinə bilməz!', 'danger')
            return redirect(url_for('employees'))
        conn.execute('DELETE FROM employees WHERE id = ?', (id,))
        audit_log(conn, session['user_id'], f'İşçi silindi: {emp["username"]}')
        conn.commit()
        flash('İşçi silindi.', 'info')
    finally:
        conn.close()
    return redirect(url_for('employees'))
@app.route('/audit')
@login_required
@role_required('Admin')
def audit():
    # Yeni Filter Parametrlərini Qəbul Et
    search_query = request.args.get('search_query', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()

    # Get page number from request, default to 1
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Number of items per page
    
    conn = get_db_connection()
    total_logs = 0
    total_pages = 0
    logs = []
    
    # 1. WHERE şərtlərini və parametrlərini dinamik qur
    where_conditions = []
    params = []

    # Axtarış sorğusu: İstifadəçi adı VƏ ya Əməliyyat adı üzrə
    if search_query:
        search_term = f"%{search_query}%"
        where_conditions.append("""(
            e.username LIKE ? OR
            a.action LIKE ?
        )""")
        params.extend([search_term, search_term])
    
    # Tarix Filteri: Başlanğıc tarixi
    if start_date:
        where_conditions.append("DATE(a.timestamp) >= ?")
        params.append(start_date)
        
    # Tarix Filteri: Bitiş tarixi
    if end_date:
        where_conditions.append("DATE(a.timestamp) <= ?")
        params.append(end_date)
        
    # WHERE hissəsini birləşdir
    where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""

    try:
        # 2. Cəmi Qeyd Sayını Hesabla (Filterləri nəzərə alaraq)
        count_query = f'''
            SELECT COUNT(a.id)
            FROM audit_log a 
            LEFT JOIN employees e ON a.employee_id = e.id
            {where_clause}
        '''
        total_logs = conn.execute(count_query, params).fetchone()[0]
        total_pages = (total_logs + per_page - 1) // per_page  # Ceiling division
        
        # Ensure page is within valid range
        page = max(1, min(page, total_pages)) if total_logs > 0 else 1
        
        # Calculate offset for pagination
        offset = (page - 1) * per_page
        
        # 3. Əsas Logları Çək (Filter, Limit və Offset ilə)
        logs_query = f'''
            SELECT 
                a.id, 
                e.username, 
                a.action, 
                a.details,
                a.timestamp
            FROM 
                audit_log a 
            LEFT JOIN 
                employees e ON a.employee_id = e.id
            {where_clause}
            ORDER BY 
                a.id DESC
            LIMIT ? OFFSET ?
        '''
        # Parametrlərin sonuna LIMIT və OFFSET dəyərlərini əlavə et
        final_params = params + [per_page, offset]
        logs = conn.execute(logs_query, final_params).fetchall()
        
    finally:
        conn.close()
        
    # Yeni axtarış dəyərlərini şablona göndər
    return render_template('audit.html', 
                         logs=logs,
                         page=page,
                         total_pages=total_pages,
                         total_logs=total_logs,
                         search_query=search_query,
                         start_date=start_date,
                         end_date=end_date)

@app.route('/products', methods=['GET', 'POST'])
@login_required
def products():
    if session['role'] not in ['Admin', 'Operator']:
        flash('Bu səhifəyə giriş icazəniz yoxdur.', 'danger')
        return redirect(url_for('index'))

    conn = get_db_connection()
    warning_items = []  

    try:
        if request.method == 'POST':
            name = request.form.get('name')
            sku = request.form.get('sku')
            category = request.form.get('category')
            unit = request.form.get('unit')
            try:
                price = float(request.form.get('price', 0))
                stock = float(request.form.get('stock', 0))
                min_stock = float(request.form.get('min_stock', 0))
            except ValueError:
                flash('Qiymət və stok rəqəm olmalıdır.', 'danger')
                return redirect(url_for('products'))

            location = request.form.get('location')
            description = request.form.get('description')
            supplier_id = request.form.get('supplier_id') or None 
            if supplier_id:
                supplier_id = int(supplier_id)
            
            try:
                conn.execute(
                    '''INSERT INTO products (name, sku, category, unit, price, stock, min_stock, location, description, supplier_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (name, sku, category, unit, price, stock, min_stock, location, description, supplier_id)
                )
                audit_log(conn, session['user_id'], f'Məhsul əlavə edildi: {name} (SKU:{sku}). İlkin Stok: {stock} {unit}')
                conn.commit()
                flash('Məhsul əlavə edildi.', 'success')
            except sqlite3.IntegrityError as e:
                if 'UNIQUE constraint failed: products.sku' in str(e):
                    flash('SKU artıq mövcuddur.', 'danger')
                elif 'FOREIGN KEY constraint failed' in str(e):
                    flash('Seçilmiş təchizatçı tapılmadı.', 'danger')
                else:
                    flash(f'Xəta baş verdi: {e}', 'danger')
            except Exception as e:
                flash(f'Xəta baş verdi: {e}', 'danger')

        cursor = conn.cursor()
        cursor.execute(
            "SELECT name, stock, unit FROM products WHERE stock < 100 AND is_active = 1 ORDER BY name"
        )
        low_stock_products = cursor.fetchall()
        
        for item in low_stock_products:
            warning_items.append({
                'name': item['name'],
                'stock': item['stock'],
                'unit': item['unit']
            })

        products_rows = conn.execute(
            'SELECT p.*, s.name AS supplier_name FROM products p LEFT JOIN suppliers s ON p.supplier_id = s.id WHERE p.is_active = 1 ORDER BY p.name'
        ).fetchall()
        
        suppliers_rows = conn.execute(
            'SELECT id, name FROM suppliers WHERE is_active = 1'
        ).fetchall()

        products_list = [dict(row) for row in products_rows]
        suppliers_list = [dict(row) for row in suppliers_rows]

    finally:
        conn.close()
        
    return render_template(
        'products.html', 
        products=products_list, 
        suppliers=suppliers_list,
        warning_items=warning_items
    )

@app.route('/delete_product/<int:id>', methods=['POST', 'GET'])
@login_required
@role_required('Operator')
def delete_product(id):
    conn = get_db_connection()
    try:
        product = conn.execute('SELECT * FROM products WHERE id = ?', (id,)).fetchone()
        if not product:
            flash('Məhsul tapılmadı.', 'danger')
            return redirect(url_for('products'))
        conn.execute('DELETE FROM products WHERE id = ?', (id,))
        audit_log(conn, session['user_id'], f'Məhsul silindi: {product["name"]} (SKU:{product["sku"]})')
        conn.commit()
        flash('Məhsul silindi.', 'success')
    except Exception as e:
        flash(f'Xəta: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('products'))


@app.route('/update_stock/<int:id>', methods=['GET', 'POST'])
@login_required
def update_stock(id):
    conn = get_db_connection()
    product = None
    try:
        product = conn.execute('SELECT * FROM products WHERE id = ?', (id,)).fetchone()
        if not product:
            flash('Məhsul tapılmadı.', 'danger')
            return redirect(url_for('products'))

        if request.method == 'POST':
            action_type = request.form.get('action_type')
            qty_change_str = request.form.get('qty_change')

            if not qty_change_str or not action_type:
                flash('Əməliyyat növü və Miqdar boş ola bilməz.', 'danger')
                return render_template('update_stock.html', product=product)

            try:
                qty_change = float(qty_change_str)
                if qty_change <= 0:
                    raise ValueError('Miqdar 0-dan böyük olmalıdır.')
            except ValueError:
                flash('Miqdar düzgün müsbət tam ədəd olmalıdır.', 'danger')
                return render_template('update_stock.html', product=product)

            current_stock = product['stock']

            if action_type == 'in':
                new_stock = current_stock + qty_change
                log_action = f'Stok Girişi (+{qty_change}): {product["name"]} (Yeni Stok: {new_stock})'
                flash_message = f'Stok əlavə edildi: +{qty_change}'
            elif action_type == 'out':
                if qty_change > current_stock:
                    flash(f'Stok çıxışı mümkün deyil! Anbarda cəmi {current_stock} ədəd var.', 'danger')
                    return render_template('update_stock.html', product=product)
                new_stock = current_stock - qty_change
                log_action = f'Stok Çıxışı (-{qty_change}): {product["name"]} (Yeni Stok: {new_stock})'
                flash_message = f'Stokdan çıxarıldı: -{qty_change}'
            else:
                flash('Yanlış əməliyyat növü seçilib.', 'danger')
                return render_template('update_stock.html', product=product)

            conn.execute('UPDATE products SET stock = ?, updated_at = ? WHERE id = ?', (new_stock, get_current_baku_time(), id))
            record_transaction(conn, id, 'in' if action_type == 'in' else 'out', qty_change, product['price'] or 0, (product['price'] or 0) * qty_change, session['user_id'])
            audit_log(conn, session['user_id'], log_action)
            conn.commit()
            flash(flash_message, 'success')
            return redirect(url_for('products'))

    finally:
        conn.close()
    return render_template('update_stock.html', product=product)
import json
import json
import time # time.time() üçün əlavə edilib
import uuid

def generate_unique_sku(prefix='PROD'):
    """Generate a unique SKU with the given prefix and a random UUID.
    
    Args:
        prefix (str): Prefix for the SKU (default: 'PROD')
        
    Returns:
        str: A unique SKU string
    """
    # Use UUID4 and take first 8 chars for shorter but still unique SKU
    # Add timestamp to ensure uniqueness even if UUID collides (very rare)
    timestamp = int(time.time() * 1000) % 10000  # Last 4 digits of timestamp
    unique_id = f"{prefix}-{str(uuid.uuid4().hex[:6])}{timestamp:04d}"
    return unique_id.upper()

@app.route('/transactions', methods=['GET', 'POST'])
@login_required
def transactions(): 
    # Giriş icazəsi yoxlanılır
    if session.get('role') not in ['Admin', 'Operator']:
        flash('Bu səhifəyə giriş icazəniz yoxdur.', 'danger')
        return redirect(url_for('index'))

    # ==============================================================================
    # GET REQUEST: Formu göstərmək və məlumatları yükləmək (Dəyişməyib)
    # ==============================================================================
    if request.method == 'GET':
        try:
            with DatabaseConnection() as conn:
                products_data = conn.execute("""
                    SELECT id, name, sku, price, stock, category, unit, location, min_stock 
                    FROM products 
                    WHERE is_active = 1 
                    ORDER BY name
                """).fetchall()
                
                suppliers_data = conn.execute("""
                    SELECT id, name 
                    FROM suppliers 
                    WHERE is_active = 1 
                    ORDER BY name
                """).fetchall()
                
                # Məlumatları sözlüklərə çevirmə
                products_list = []
                for row in products_data:
                    product = dict(row)
                    
                    product['id'] = int(product['id']) if product['id'] is not None else 0
                    product['price'] = float(product['price']) if product['price'] is not None else 0.0
                    product['stock'] = float(product['stock']) if product['stock'] is not None else 0.0
                    product['min_stock'] = float(product['min_stock']) if product['min_stock'] is not None else 0.0
                    
                    products_list.append(product) 

                suppliers = [dict(row) for row in suppliers_data]
                customers = [{'name': 'Anonim Müştəri'}] 
                
                return render_template('transactions.html', 
                                       products=products_list, 
                                       suppliers=suppliers,
                                       customers=customers)
        except Exception as e:
            app.logger.error(f'GET Transaction Error: {e}', exc_info=True)
            flash('Məlumatların yüklənməsi zamanı xəta baş verdi.', 'danger')
            return redirect(url_for('index'))


    # ==============================================================================
    # POST REQUEST: Əməliyyatı emal etmək
    # ==============================================================================
    with DatabaseConnection() as conn:
        try:
            type_ = request.form.get('transaction_type') 
            employee_id = session.get('user_id', 1) 

            # 1. Frontend-dən gələn məhsul datasını JSON formatında oxumaq
            products_json_key = f'{type_}_products_json' 
            products_data_json = request.form.get(products_json_key)
            
            if not products_data_json:
                flash('Əməliyyat üçün məhsul seçilməyib!', 'danger')
                return redirect(url_for('transactions'))
            
            try:
                selected_products = json.loads(products_data_json)
                if not isinstance(selected_products, list) or not selected_products:
                    # Əlavə yoxlama: siyahı boş ola bilməz
                    raise ValueError("Empty or invalid products data format")
            except (json.JSONDecodeError, ValueError) as e:
                app.logger.error(f'JSON parse error: {e}', exc_info=True)
                flash('Məhsul məlumatları düzgün formatda deyil!', 'danger')
                return redirect(url_for('transactions'))

            # 2. Ümumi Məlumatları oxumaq
            total = float(request.form.get('total', 0))
            paid_amount = float(request.form.get('paid_amount', 0))
            # 🔴 DÜZƏLİŞ: .strip() əlavə edildi ki, boşluqları təmizləsin.
            note = request.form.get(f'{type_}_note', '').strip()
            
            # Baza Validasiyası
            if paid_amount < 0:
                flash('Ödənilən məbləğ mənfi ola bilməz!', 'danger')
                return redirect(url_for('transactions'))
            
            # ==============================================================================
            # ALMA ƏMƏLİYYATI (PURCHASE)
            # ==============================================================================
            if type_ == 'purchase':
                supplier_id_str = request.form.get('supplier_id')
                if not supplier_id_str or not supplier_id_str.isdigit():
                    flash('Təchizatçı seçmək məcburidir!', 'danger')
                    return redirect(url_for('transactions'))
                
                supplier_id = int(supplier_id_str)
                supplier = conn.execute("SELECT id, name FROM suppliers WHERE id = ? AND is_active = 1", (supplier_id,)).fetchone()
                if not supplier:
                    flash('Seçilmiş təchizatçı tapılmadı və ya aktiv deyil!', 'danger')
                    return redirect(url_for('transactions'))
                    
                # Ödəniş Validasiyası
                if paid_amount > total and total > 0: 
                    flash('Ödənilən məbləğ ümumi məbləğdən çox ola bilməz!', 'danger')
                    return redirect(url_for('transactions'))
                
                # Hər bir məhsul üçün loop
                processed_products = []
                total_purchase_amount = 0.0
                
                for item in selected_products:
                    # DÜZƏLİŞ BAŞLANĞICI
                    product_id_str = str(item.get('id', ''))
                    product_id = None 
                    is_new = False
                    
                    if product_id_str.startswith('new_'):
                        # 1. Yeni Məhsul
                        is_new = True
                        product_id = None 
                    else:
                        # 2. Mövcud Məhsul
                        try:
                            product_id = int(product_id_str) 
                            if product_id <= 0:
                                raise ValueError("Invalid product ID: Must be positive.") 
                        except (ValueError, TypeError) as e:
                            app.logger.error(f"Invalid product ID format for existing product: {product_id_str}, error: {e}")
                            flash(f'Mövcud məhsul üçün yanlış ID formatı: {product_id_str}', 'danger')
                            return redirect(url_for('transactions'))
                    # DÜZƏLİŞ SONU
                            
                    quantity = float(item.get('quantity', 0))
                    price = float(item.get('price', 0)) 
                    
                    if quantity <= 0 or price < 0: 
                        continue 
                    
                    total_item_cost = quantity * price
                    total_purchase_amount += total_item_cost
                    
                    # Log üçün məhsul detallarını hazırlayırıq
                    product_data = {
                        'is_new': is_new,
                        'quantity': quantity,
                        'price': price,
                        'total': total_item_cost,
                        'note': item.get('note', note),
                        'product_name': item.get('product_name', 'Naməlum Məhsul'),
                        'unit': item.get('unit', 'ədəd'),
                        'category': item.get('category', 'Digər'),
                        'id': product_id 
                    }
                    
                    processed_products.append(product_data)
                
                # Həqiqi ümumi məbləği yenidən hesablayırıq
                total = total_purchase_amount
                if total <= 0:
                    flash('Əməliyyatın ümumi məbləği sıfırdır. Məhsul daxil edilib, miqdar və ya qiymət yoxlanılmalıdır.', 'danger')
                    return redirect(url_for('transactions'))
                
                # Ödəniş Validasiyası (Total yeniləndikdən sonra yenidən yoxlama)
                if paid_amount > total:
                    flash(f'Ödənilən məbləğ ({paid_amount:.2f}) ümumi məbləğdən ({total:.2f}) çox ola bilməz!', 'danger')
                    return redirect(url_for('transactions'))

                
                # Hər bir məhsulu emal et (Stok artımı və Tranzaksiya qeydi)
                for item in processed_products:
                    try:
                        if item['is_new']:
                            # YENİ MƏHSUL YARADILMASI MƏNTİQİ
                            try:
                                # 1. Məhsulu yarat
                                cursor = conn.execute("""
                                    INSERT INTO products (name, sku, price, stock, category, unit, location, min_stock, supplier_id, is_active, created_at)
                                    VALUES (?, ?, ?, ?, ?, ?, 'Anbar', 0, ?, 1, ?)
                                """, (
                                    item['product_name'], 
                                    generate_unique_sku(), 
                                    item['price'], # Satış qiyməti olaraq ilkin alış qiyməti (sonradan dəyişdirilməlidir)
                                    item['quantity'], 
                                    item['category'], 
                                    item['unit'], 
                                    supplier_id, 
                                    get_current_baku_time()
                                ))
                                product_id = cursor.lastrowid
                                
                                # 2. Tranzaksiyanı qeyd et
                                record_transaction(
                                    conn, product_id, 'purchase', item['quantity'], 
                                    item['price'], item['total'], employee_id, item['note'], 
                                    customer_name='', supplier_id=supplier_id
                                )
                                
                                # YENİ MƏHSULUN YARADILMASINI AYRI LOG OLARAQ SAXLAYIRIQ (vacib hadisədir)
                                audit_log(conn, employee_id, 'Yeni məhsul yaradıldı', {
                                    'type': 'new_product', 'product_id': product_id, 'product_name': item['product_name'],
                                })
                                
                            except Exception as e:
                                app.logger.error(f"Yeni məhsul yaratma xətası: {e}")
                                raise
                                
                        else:
                            # MÖVCUD MƏHSULUN YENİLƏNMƏSİ
                            product_id = item['id']
                            
                            # Mövcud məhsulun məlumatlarını yüklə
                            product = conn.execute("""
                                SELECT id, name, stock, price 
                                FROM products 
                                WHERE id = ? AND is_active = 1
                            """, (product_id,)).fetchone()
                            
                            if not product:
                                raise ValueError(f"Məhsul tapılmadı: ID {product_id}")
                            
                            # Stok və WAC hesablamaları
                            old_stock = product['stock']
                            old_cost = product['price'] 
                            
                            # Yeni WAC hesablanması
                            new_total_stock = old_stock + item['quantity']
                            if new_total_stock > 0 and old_stock > 0:
                                new_wac = ((old_stock * old_cost) + (item['quantity'] * item['price'])) / new_total_stock
                            else:
                                new_wac = item['price']
                            
                            # Məhsul məlumatlarını yenilə
                            conn.execute("""
                                UPDATE products 
                                SET stock = stock + ?, 
                                    price = ?, 
                                    updated_at = ? 
                                WHERE id = ?
                            """, (
                                item['quantity'], 
                                new_wac, 
                                get_current_baku_time(), 
                                product_id
                            ))
                            
                            # Tranzaksiyanı qeyd et
                            record_transaction(
                                conn, product_id, 'purchase', item['quantity'], 
                                item['price'], item['total'], employee_id, item['note'], 
                                customer_name='', supplier_id=supplier_id
                            )
                            
                            # KONSOLİDASİYA ÜÇÜN: Burada olan 'audit_log' silindi.
                            
                    except Exception as e:
                        app.logger.error(f"Məhsul emalı xətası (ID: {item.get('id', 'yeni')}): {e}")
                        conn.rollback()
                        flash(f"Xəta baş verdi: {str(e)}. Əməliyyat geri qaytarıldı.", 'danger')
                        return redirect(url_for('transactions'))
                            
                
                # Ümumi Maliyyə, Borc VƏ KONSOLİDASİYA EDİLMİŞ AUDIT LOG Məntiqi
                try:
                    supplier_name_row = conn.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,)).fetchone()
                    supplier_name = supplier_name_row['name'] if supplier_name_row else 'Naməlum Təchizatçı'
                    
                    # Ödəniş qeydi
                    if paid_amount > 0:
                        conn.execute('''
                            INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id, supplier_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', ('Expense', 'Məhsul Alışı', paid_amount, f'Çoxlu məhsul alışı üçün ödəniş. Təchizatçı ID: {supplier_id}', get_current_baku_time(), employee_id, supplier_id))
                        # KONSOLİDASİYA ÜÇÜN: Əvvəlki audit_log burdan silindi.

                    remaining_amount = total - paid_amount 
                    
                    # Borc qeydi
                    if remaining_amount > 0.01: 
                        conn.execute('''
                            INSERT INTO debts (party_name, type, initial_amount, paid_amount, description, status, created_at, employee_id, supplier_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (supplier_name, 'Debt', total, paid_amount, f'Çoxlu məhsul alınması zamanı yaranan borc. Təchizatçı: {supplier_name}', 'Gözlənilir', get_current_baku_time(), employee_id, supplier_id))
                        # KONSOLİDASİYA ÜÇÜN: Əvvəlki audit_log burdan silindi.
                    
                    
                    # ******************************************************************************
                    # YENİ: ƏSAS ALMA ƏMƏLİYYATINI KONSOLİDASİYA EDƏN AUDIT LOG-UN YARADILMASI
                    # ******************************************************************************
                    audit_details = {
                        'action_type': 'purchase_completed',
                        'supplier_id': supplier_id,
                        'supplier_name': supplier_name, 
                        'total_amount': total,
                        'paid_amount': paid_amount,
                        'remaining_debt': remaining_amount,
                        'note': note,
                        # Bütün məhsul detalları
                        'products': processed_products 
                    }
                    
                    action_description = f'Əsas ALMA Əməliyyatı tamamlandı. Ümumi: {total:.2f} AZN. Ödəniş: {paid_amount:.2f} AZN'

                    audit_log(
                        conn, 
                        employee_id, 
                        action_description, 
                        audit_details
                    )
                    # ******************************************************************************

                    conn.commit()
                    success_msg = 'Alış əməliyyatı uğurla tamamlandı!'
                    if remaining_amount > 0:
                        success_msg += f' Qalıq borc: {remaining_amount:.2f} AZN'
                    flash(success_msg, 'success')
                    
                except Exception as e:
                    conn.rollback()
                    app.logger.error(f'Maliyyə əməliyyatları zamanı xəta: {e}')
                    flash('Maliyyə əməliyyatları zamanı xəta baş verdi. Əməliyyat geri qaytarıldı.', 'danger')
                    return redirect(url_for('transactions'))
                    
                return redirect(url_for('transactions'))
                
            
            # ==============================================================================
            # SATIŞ ƏMƏLİYYATI (SALE)
            # ==============================================================================
            elif type_ == 'sale':
                customer_name = request.form.get('customer_name', 'Müştəri').strip() or 'Müştəri'
                
                products_to_sell = []
                total_sale_amount = 0
                
                # 1. Bütün məhsullar üçün stok yoxlaması
                for item in selected_products:
                    # Satışda yeni məhsul satışı olmadığı üçün ID-nin int olması şərtdir.
                    try:
                        product_id = int(item.get('id', 0))
                        if product_id <= 0:
                            raise ValueError("Invalid product ID")
                    except (ValueError, TypeError):
                        flash('Satış üçün yanlış məhsul ID formatı!', 'danger')
                        return redirect(url_for('transactions'))

                    quantity = float(item.get('quantity', 0))
                    price = float(item.get('price', 0))
                    
                    if quantity <= 0 or price < 0:
                        continue
                        
                    product = conn.execute('''
                        SELECT id, name, stock, price as cost_price 
                        FROM products 
                        WHERE id = ? AND is_active = 1
                    ''', (product_id,)).fetchone()
                    
                    if not product:
                        flash(f'ID-si {product_id} olan məhsul tapılmadı və ya aktiv deyil!', 'danger')
                        return redirect(url_for('transactions'))
                    
                    if product['stock'] < quantity:
                        flash(f'"{product["name"]}" üçün kifayət qədər stok yoxdur! Anbarda: {product["stock"]}', 'danger')
                        return redirect(url_for('transactions'))
                    
                    total_item_amount = quantity * price
                    total_sale_amount += total_item_amount
                    
                    # Log üçün bütün məlumatları saxlayırıq
                    products_to_sell.append({
                        'id': product_id, 'name': product['name'], 'quantity': quantity, 
                        'price': price, 'cost_price': product['cost_price'], 
                        'total': total_item_amount, 'note': item.get('note', note)
                    })

                total = total_sale_amount
                if total <= 0:
                    flash('Satışın ümumi məbləği sıfırdır.', 'danger')
                    return redirect(url_for('transactions'))
                
                try:
                    # 2. Bütün məhsulları emal et
                    for item in products_to_sell:
                        # Stoku azalt, Tranzaksiya qeyd et
                        conn.execute('''UPDATE products SET stock = stock - ?, updated_at = ? WHERE id = ?''', (item['quantity'], get_current_baku_time(), item['id']))
                        record_transaction(conn, item['id'], 'sale', item['quantity'], item['price'], item['total'], employee_id, item['note'], customer_name)
                        
                        # KONSOLİDASİYA ÜÇÜN: Əvvəlki audit_log burdan silindi.
                    
                    # 3. Ödəniş və borc qeydləri
                    if paid_amount > 0:
                        conn.execute('''
                            INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id, customer_name) 
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', ('Income', 'Məhsul Satışı', paid_amount, f'Müştəridən alınan ödəniş: {customer_name}', get_current_baku_time(), employee_id, customer_name))
                    
                    remaining_amount = total_sale_amount - paid_amount
                    if remaining_amount > 0.01:
                        conn.execute('''
                            INSERT INTO debts (party_name, type, initial_amount, paid_amount, description, status, created_at, employee_id) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (customer_name, 'Receivable', total_sale_amount, paid_amount, f'Məhsul satışı üzrə qalıq borc. Müştəri: {customer_name}', 'Gözlənilir', get_current_baku_time(), employee_id))
                    
                    
                    # ******************************************************************************
                    # YENİ: ƏSAS SATIŞ ƏMƏLİYYATINI KONSOLİDASİYA EDƏN AUDIT LOG-UN YARADILMASI
                    # ******************************************************************************
                    total_profit = sum((item['price'] - item['cost_price']) * item['quantity'] for item in products_to_sell)

                    audit_details = {
                        'action_type': 'sale_completed',
                        'customer_name': customer_name,
                        'total_amount': total_sale_amount, 
                        'paid_amount': paid_amount,
                        'remaining_receivable': remaining_amount,
                        'total_profit': total_profit, 
                        'note': note,
                        # Bütün məhsul detalları
                        'products': products_to_sell 
                    }

                    action_description = f'Əsas SATIŞ Əməliyyatı tamamlandı. Ümumi: {total_sale_amount:.2f} AZN. Ödəniş: {paid_amount:.2f} AZN'

                    audit_log(
                        conn, 
                        employee_id, 
                        action_description, 
                        audit_details
                    )
                    # ******************************************************************************
                    
                    # Təsdiqlə
                    conn.commit()
                    
                    success_msg = f'Satış uğurla tamamlandı! Ümumi məbləğ: {total_sale_amount:.2f} AZN'
                    if remaining_amount > 0:
                        success_msg += f', Qalıq borc: {remaining_amount:.2f} AZN'
                    
                    flash(success_msg, 'success')
                    
                except Exception as e:
                    conn.rollback()
                    app.logger.error(f'Satış əməliyyatı zamanı xəta: {e}', exc_info=True)
                    flash('Satış əməliyyatı zamanı xəta baş verdi. Dəyişikliklər geri qaytarıldı.', 'danger')
                    return redirect(url_for('transactions'))
                
                return redirect(url_for('transactions'))

            # ==============================================================================
            # YANLIŞ ƏMƏLİYYAT (Dəyişməyib)
            # ==============================================================================
            else:
                flash('Yanlış əməliyyat növü!', 'danger')
                return redirect(url_for('transactions'))

        except Exception as e:
            conn.rollback()
            # Xətanın növünü dəqiqləşdiririk
            if isinstance(e, sqlite3.IntegrityError):
                log_message = f'SQLite Integrity Error: {e}'
                user_message = f'Verilənlər bazası xətası (Təkrarlanan dəyər və ya Məcburi Sahə): {e}'
            else:
                log_message = f'Transaction error: {e}'
                user_message = f'Əməliyyat zamanı kritik xəta baş verdi. Zəhmət olmasa loglara baxın.'
                
            app.logger.error(log_message, exc_info=True)
            flash(user_message, 'danger')
            return redirect(url_for('transactions'))
@app.route('/finance/dashboard')
@login_required 
@role_required('Admin')
def finance_dashboard():
    conn = get_db_connection()
    try:
        income_row = conn.execute("SELECT SUM(amount) AS total FROM financial_transactions WHERE type = 'Income'").fetchone()
        expense_row = conn.execute("SELECT SUM(amount) AS total FROM financial_transactions WHERE type = 'Expense'").fetchone()
        total_income = income_row['total'] or 0
        total_expense = expense_row['total'] or 0
        net_balance = total_income - total_expense

        debts_status = conn.execute('''
            SELECT 
                SUM(CASE WHEN type IN ('receivable', 'Receivable') THEN (initial_amount - paid_amount) ELSE 0 END) AS total_receivables, 
                SUM(CASE WHEN type IN ('debt', 'Debt') THEN (initial_amount - paid_amount) ELSE 0 END) AS total_debts 
            FROM debts 
            WHERE (initial_amount - paid_amount) > 0.01
        ''').fetchone()
        
        customer_debts_status = conn.execute('''
            SELECT SUM(amount - paid_amount) AS customer_receivables 
            FROM customer_debts 
            WHERE status != 'Ödənilib' AND (amount - paid_amount) > 0.01
        ''').fetchone()
        
        total_receivables = (debts_status['total_receivables'] or 0) + (customer_debts_status['customer_receivables'] or 0)
        total_debts = (debts_status['total_debts'] or 0)

        latest_transactions = conn.execute('SELECT id, type, category, amount, description, timestamp, employee_id FROM financial_transactions ORDER BY timestamp DESC LIMIT 10').fetchall()
        
        processed_transactions = []
        for t in latest_transactions:
            t_dict = dict(t)
            try:
                t_dict['timestamp'] = datetime.datetime.strptime(t['timestamp'], '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                pass 
            processed_transactions.append(t_dict)
        
        return render_template('finance_dashboard.html', 
            total_income=total_income, 
            total_expense=total_expense, 
            net_balance=net_balance, 
            total_receivables=total_receivables, 
            total_debts=total_debts, 
            latest_transactions=processed_transactions 
        )

    except Exception as e:
        print(f"Error loading dashboard: {str(e)}")
        flash('Dashboard yüklənərkən xəta baş verdi: ' + str(e), 'danger')
        return render_template('finance_dashboard.html', 
            total_income=0, total_expense=0, net_balance=0, 
            total_receivables=0, total_debts=0, latest_transactions=[]
        )
    finally:
        conn.close()
from urllib.parse import quote
from io import BytesIO

@app.route('/finance/export', methods=['POST'])
@login_required
@role_required('Admin')
def export_financial_transactions():
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        products = cursor.execute("SELECT id, name, price FROM products").fetchall()
        transactions = cursor.execute("""
            SELECT product_id, type, quantity, price, timestamp
            FROM transactions
            WHERE product_id IS NOT NULL
        """).fetchall()

        product_summary = []

        for p in products:
            pid = p['id']
            pname = p['name']
            avg_purchase_price = p['price'] or 0

            sales = [t for t in transactions if t['product_id'] == pid and t['type'].lower() == 'sale']
            purchases = [t for t in transactions if t['product_id'] == pid and t['type'].lower() in ['purchase','in']]

            total_sales_qty = sum(t['quantity'] for t in sales)
            total_sales_amount = sum(t['price']*t['quantity'] for t in sales) if sales else 0
            avg_sales_price = (total_sales_amount / total_sales_qty) if total_sales_qty > 0 else 0

            total_purchase_qty = sum(t['quantity'] for t in purchases)
            total_purchase_amount = sum(t['price']*t['quantity'] for t in purchases) if purchases else 0

            gross_profit = total_sales_amount - (avg_purchase_price * total_sales_qty)
            profit_per_unit = (gross_profit / total_sales_qty) if total_sales_qty > 0 else 0

            product_summary.append({
                'product_id': pid,
                'product_name': pname,
                'Satış_Sayı': total_sales_qty,
                'Satış_Gəliri (AZN)': total_sales_amount,
                'Ortalama_Satış_Qiyməti (AZN)': avg_sales_price,
                'Ortalama_Alış_Qiyməti (AZN)': avg_purchase_price,
                'Alış_Xərci (AZN)': avg_purchase_price * total_sales_qty,
                'Xalis_Mənfəət (AZN)': gross_profit,
                'Hər_Satış_Qazancı (AZN)': profit_per_unit
            })

        df_products = pd.DataFrame(product_summary)

        product_totals = pd.DataFrame([{
            'product_id': '',
            'product_name': 'Ümumi',
            'Satış_Sayı': df_products['Satış_Sayı'].sum(),
            'Satış_Gəliri (AZN)': df_products['Satış_Gəliri (AZN)'].sum(),
            'Ortalama_Satış_Qiyməti (AZN)': '',
            'Ortalama_Alış_Qiyməti (AZN)': '',
            'Alış_Xərci (AZN)': df_products['Alış_Xərci (AZN)'].sum(),
            'Xalis_Mənfəət (AZN)': df_products['Xalis_Mənfəət (AZN)'].sum(),
            'Hər_Satış_Qazancı (AZN)': ''
        }])
        df_products = pd.concat([product_totals, df_products], ignore_index=True)

        cash_transactions = cursor.execute("""
            SELECT id, type, description, amount, timestamp
            FROM financial_transactions
        """).fetchall()

        df_cash = pd.DataFrame([{
            'ID': t['id'],
            'Tip': t['type'],
            'Təsvir': t['description'],
            'Məbləğ (AZN)': t['amount'],
            'Tarix': t['timestamp']
        } for t in cash_transactions])

        total_income = df_cash[df_cash['Tip'].str.lower() == 'income']['Məbləğ (AZN)'].sum()
        total_expense = df_cash[df_cash['Tip'].str.lower() == 'expense']['Məbləğ (AZN)'].sum()
        totals_row = pd.DataFrame([{
            'ID': '',
            'Tip': 'Ümumi',
            'Təsvir': '',
            'Məbləğ (AZN)': f'Gelir: {total_income} / Xərc: {total_expense}',
            'Tarix': ''
        }])
        df_cash = pd.concat([totals_row, df_cash], ignore_index=True)

   
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_products.to_excel(writer, index=False, sheet_name='Məhsul Əsaslı Təhlil')
            df_cash.to_excel(writer, index=False, sheet_name='Medaxil & Mexaric')

        output.seek(0)
        file_name = f"Maliyyə_Hesabatı_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_name_header = quote(file_name)

        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{file_name_header}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        flash('Maliyyə hesabatı uğurla eksport edildi.', 'success')
        return response

    except Exception as e:
        flash(f'Maliyyə hesabatı zamanı xəta baş verdi: {e}', 'danger')
        app.logger.error(f'Finance export error: {e}', exc_info=True)
        return redirect(url_for('finance_dashboard'))

    finally:
        conn.close()

@app.route('/finance/export/daily-earnings')
@login_required
@role_required('Admin')
def export_daily_earnings():
    """Export daily earnings to Excel with detailed transaction information"""
    try:
        # Get the current date in Baku timezone
        baku_tz = pytz.timezone('Asia/Baku')
        today = datetime.datetime.now(baku_tz).date()
        
        with DatabaseConnection() as conn:
            # Query to get daily summary
            summary_query = """
            SELECT 
                strftime('%Y-%m-%d', t.timestamp) as date,
                COALESCE(SUM(CASE WHEN t.type = 'sale' THEN t.total ELSE 0 END), 0) as total_sales,
                COALESCE(SUM(CASE WHEN t.type = 'purchase' THEN t.total ELSE 0 END), 0) as total_expenses,
                COALESCE(SUM(CASE WHEN t.type = 'sale' THEN t.total ELSE 0 END), 0) - 
                COALESCE(SUM(CASE WHEN t.type = 'purchase' THEN t.total ELSE 0 END), 0) as net_earnings
            FROM transactions t
            WHERE date(t.timestamp) = date(?)
            GROUP BY date(t.timestamp)
            """
            
            # Query to get detailed transactions
            details_query = """
            SELECT 
                t.id,
                p.name as product_name,
                t.type,
                t.quantity,
                t.price,
                t.total,
                t.timestamp,
                e.username as employee_username,
                CASE 
                    WHEN t.type = 'sale' THEN t.customer_name
                    WHEN t.type = 'purchase' THEN s.name
                    ELSE 'N/A'
                END as party_name
            FROM transactions t
            LEFT JOIN products p ON t.product_id = p.id
            LEFT JOIN employees e ON t.employee_id = e.id
            LEFT JOIN suppliers s ON t.supplier_id = s.id
            WHERE date(t.timestamp) = date(?)
            ORDER BY t.timestamp DESC
            """
            
            # Execute queries with today's date
            today_str = today.strftime('%Y-%m-%d')
            
            # Get summary data
            cursor = conn.execute(summary_query, (today_str,))
            summary = cursor.fetchone()
            
            # Get detailed transactions
            cursor = conn.execute(details_query, (today_str,))
            transactions = cursor.fetchall()
            
            if not transactions:
                flash('Bu gün üçün məlumat tapılmadı', 'warning')
                return redirect(url_for('finance_dashboard'))
            
            # Create Excel file in memory
            output = BytesIO()
            wb = openpyxl.Workbook()
            
            # Add Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Ümumi Məlumat"
            
            # Add summary headers
            summary_headers = [
                "Tarix", 
                "Ümumi Satış (AZN)", 
                "Ümumi Xərc (AZN)", 
                "Xalis Qazanc (AZN)"
            ]
            ws_summary.append(summary_headers)
            
            # Add summary data
            summary_row = [
                summary['date'],
                round(float(summary['total_sales'] or 0), 2),
                round(float(summary['total_expenses'] or 0), 2),
                round(float(summary['net_earnings'] or 0), 2)
            ]
            ws_summary.append(summary_row)
            
            # Format summary sheet
            for col in ws_summary.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws_summary.column_dimensions[column].width = adjusted_width
            
            # Add Transactions sheet
            ws_transactions = wb.create_sheet("Ətraflı Hesabat")
            
            # Add transaction headers
            transaction_headers = [
                "Tarix/Saat",
                "Əməliyyat Növü",
                "Məhsul Adı",
                "Miqdar",
                "Vahid Qiymət (AZN)",
                "Ümumi Məbləğ (AZN)",
                "İşçi",
                "Tərəf"
            ]
            ws_transactions.append(transaction_headers)
            
            # Add transaction data
            for tx in transactions:
                tx_type = 'Satış' if tx['type'] == 'sale' else 'Alış'
                tx_row = [
                    tx['timestamp'],
                    tx_type,
                    tx['product_name'] or 'N/A',
                    round(float(tx['quantity'] or 0), 2),
                    round(float(tx['price'] or 0), 2),
                    round(float(tx['total'] or 0), 2),
                    tx['employee_username'] or 'N/A',
                    tx['party_name'] or 'N/A'
                ]
                ws_transactions.append(tx_row)
            
            # Format transactions sheet
            for col in ws_transactions.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws_transactions.column_dimensions[column].width = adjusted_width
            
            # Set date format for timestamp column
            for row in ws_transactions.iter_rows(min_row=2, max_row=len(transactions)+1, min_col=1, max_col=1):
                for cell in row:
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            
            # Format currency columns
            for row in ws_transactions.iter_rows(min_row=2, max_row=len(transactions)+1, min_col=4, max_col=6):
                for cell in row:
                    cell.number_format = '#,##0.00'
            
            # Format summary sheet currency
            for row in ws_summary.iter_rows(min_row=2, max_row=2, min_col=2, max_col=4):
                for cell in row:
                    cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths for both sheets
            for sheet in [ws_summary, ws_transactions]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column_letter].width = min(50, adjusted_width)  # Cap width at 50
            
            # Save the workbook to the BytesIO object
            wb.save(output)
            output.seek(0)
            
            # Create response with Excel file
            filename = f'gunluk_hesabat_{today}.xlsx'
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            flash('Günlük hesabat uğurla yükləndi', 'success')
            return response
            
    except Exception as e:
        app.logger.error(f"Error generating daily earnings report: {str(e)}")
        flash(f'Xəta baş verdi: {str(e)}', 'danger')
        return redirect(url_for('finance_dashboard'))

@app.route('/finance/transactions', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def financial_transactions():
    conn = get_db_connection()
    try:
        if request.method == 'POST':
            type_ = 'Income' if request.form.get('type') == 'Gəlir' else 'Expense'
            category = request.form.get('category')
            description = request.form.get('description')
            try:
                amount = float(request.form.get('amount'))
                if amount <= 0:
                    raise ValueError()
            except Exception:
                flash('Məbləğ düzgün rəqəm formatında olmalıdır.', 'danger')
                return redirect(url_for('financial_transactions'))

            conn.execute('BEGIN')
            try:
                conn.execute('INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id) VALUES (?, ?, ?, ?, ?, ?)', (type_, category, amount, description, get_current_baku_time(), session['user_id']))
                audit_log(conn, session['user_id'], f'Yeni maliyyə əməliyyatı: {type_} {amount}')
                conn.commit()
                flash('Əməliyyat uğurla əlavə edildi.', 'success')
            except Exception as e:
                conn.rollback()
                app.logger.error(f'Error in financial transaction: {e}')
                flash('Əməliyyat zamanı xəta baş verdi.', 'danger')
                return redirect(url_for('financial_transactions'))

        # Get search query parameters
        search_query = request.args.get('search', '').strip()
        search_type = request.args.get('search_type', 'all')
        
        # Base query and parameters
        query = 'FROM financial_transactions'
        params = []
        
        # Add search conditions if search query exists
        if search_query:
            search_conditions = []
            if search_type in ['all', 'type']:
                search_conditions.append("type LIKE ?")
                params.append(f'%{search_query}%')
            if search_type in ['all', 'category'] and search_query:
                search_conditions.append("category LIKE ?")
                params.append(f'%{search_query}%')
            if search_type in ['all', 'description'] and search_query:
                search_conditions.append("description LIKE ?")
                params.append(f'%{search_query}%')
            if search_type in ['all', 'amount'] and search_query.replace('.', '').isdigit():
                search_conditions.append("amount = ?")
                params.append(float(search_query))
            
            if search_conditions:
                query += " WHERE " + " OR ".join(search_conditions)
        
        # Get total number of transactions for pagination
        count_query = f"SELECT COUNT(*) as count {query}"
        total_transactions = conn.execute(count_query, params).fetchone()['count']
        
        # Get page number from query parameter, default to 1
        page = request.args.get('page', 1, type=int)
        per_page = 50
        total_pages = max(1, (total_transactions + per_page - 1) // per_page)
        
        # Ensure page is within valid range
        page = max(1, min(page, total_pages))
        offset = (page - 1) * per_page
        
        # Get paginated transactions
        transactions_query = f'''
            SELECT id, type, category, amount, description, timestamp, employee_id 
            {query}
            ORDER BY timestamp DESC 
            LIMIT ? OFFSET ?
        '''
        params.extend([per_page, offset])
        transactions_list = conn.execute(transactions_query, params).fetchall()
        
        processed_transactions = []
        for t in transactions_list:
            t_dict = dict(t)
            try:
                t_dict['timestamp'] = datetime.datetime.strptime(t['timestamp'], '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                pass 
            processed_transactions.append(t_dict)
            
        return render_template('financial_transactions.html', 
                             transactions=processed_transactions,
                             current_page=page,
                             total_pages=total_pages,
                             total_transactions=total_transactions)
    finally:
        conn.close()
def get_party_name(row):
    try:
        return row['party_name']
    except:
        return ''
from reportlab.lib.utils import simpleSplit 

@app.route('/invoice/download/<int:transaction_id>')
@login_required
@role_required('Admin')
def download_invoice(transaction_id):
    conn = get_db_connection()
    try:
        transaction = conn.execute(
            'SELECT * FROM financial_transactions WHERE id = ?', (transaction_id,)
        ).fetchone()

        if not transaction:
            flash('Faktura/Əməliyyat tapılmadı.', 'danger')
            return redirect(url_for('invoices'))

        employee_info = f"ID: {transaction['employee_id']}"
        description_text = transaction['description']

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        current_y = height - 3 * cm

        party_name = get_party_name(transaction)
        
        if not party_name and description_text:
            try:
                parts = description_text.split('-')
                
                if len(parts) >= 3:
                    party_name = parts[1].strip()
                
                elif 'Müştəri:' in description_text:
                    party_name = description_text.split('Müştəri:')[1].strip()
                    
            except Exception:
                pass 
        
        if not party_name:
            party_name = 'Tərəf Adı Təyin Edilməyib'
        
        c.setFont('Helvetica-Bold', 18)
        c.drawCentredString(width / 2, current_y, transliterate_az('HESAB FAKTURA'))
        c.setFont('Helvetica', 10)
        current_y -= 1.5 * cm
        
        c.setFont('Helvetica-Bold', 12)
        c.drawString(2 * cm, current_y, transliterate_az("Faktura Məlumatları"))
        c.line(2 * cm, current_y - 0.1 * cm, 8 * cm, current_y - 0.1 * cm)
        current_y -= 0.6 * cm
        
        c.setFont('Helvetica', 10)
        c.drawString(2 * cm, current_y, transliterate_az(f"Faktura ID:"))
        c.drawString(5 * cm, current_y, f"{transaction['id']}")
        current_y -= 0.5 * cm
        
        c.drawString(2 * cm, current_y, transliterate_az(f"Tarix:"))
        c.drawString(5 * cm, current_y, f"{transaction['timestamp']}")
        current_y -= 0.5 * cm

        c.drawString(2 * cm, current_y, transliterate_az(f"Əməliyyat Növü:"))
        c.drawString(5 * cm, current_y, transliterate_az(transaction['type']))
        current_y -= 1 * cm
        
        c.setFont('Helvetica-Bold', 12)
        c.drawString(2 * cm, current_y, transliterate_az("Ödəyən/Alıcı Tərəf"))
        c.line(2 * cm, current_y - 0.1 * cm, 8 * cm, current_y - 0.1 * cm)
        current_y -= 0.6 * cm
        
        c.setFont('Helvetica', 10)
        c.drawString(2 * cm, current_y, transliterate_az(f"Tərəf:"))
        c.drawString(5 * cm, current_y, transliterate_az(party_name))
        current_y -= 1 * cm
        
        c.setFont('Helvetica-Bold', 12)
        c.drawString(2 * cm, current_y, transliterate_az("Əməliyyatın Təsviri"))
        c.line(2 * cm, current_y - 0.1 * cm, 8 * cm, current_y - 0.1 * cm)
        current_y -= 0.6 * cm
        
        c.setFont('Helvetica', 10)
        
        description_text_az = transliterate_az(description_text)
        
        c.setFont('Helvetica', 9) 
        textobject = c.beginText(2 * cm, current_y)
        textobject.setLeading(12) 
        
        max_width = 17 * cm 
        
        lines = simpleSplit(description_text_az, 'Helvetica', 9, max_width)
        
        for line in lines:
            textobject.textLine(line)

        c.drawText(textobject)
        
        current_y -= (len(lines) * 0.45) * cm + 1 * cm 

        
        y_start = current_y
        c.setFont('Helvetica-Bold', 10)
        c.drawString(2 * cm, y_start, transliterate_az("Kateqoriya"))
        c.drawString(9 * cm, y_start, transliterate_az("Əməliyyatçı"))
        c.drawRightString(18 * cm, y_start, transliterate_az("Məbləğ (AZN)"))

        c.line(2 * cm, y_start - 0.2 * cm, 19 * cm, y_start - 0.2 * cm)

        y_data = y_start - 1 * cm
        c.setFont('Helvetica', 10)
        c.drawString(2 * cm, y_data, transliterate_az(transaction['category']))
        c.drawString(9 * cm, y_data, transliterate_az(employee_info))
        
        c.drawRightString(18 * cm, y_data, f"{transaction['amount']:.2f} AZN")

        c.line(2 * cm, y_data - 0.2 * cm, 19 * cm, y_data - 0.2 * cm)

        c.setFont('Helvetica-Bold', 14)
        c.drawRightString(18 * cm, y_data - 2 * cm, transliterate_az(f"Yekun Məbləğ: {transaction['amount']:.2f} AZN"))

        c.setFont('Helvetica', 10)
        c.drawString(3 * cm, 5 * cm, transliterate_az("Mühasib / Rəhbər"))
        c.drawString(12 * cm, 5 * cm, transliterate_az(f"Əməliyyatçı ({employee_info})"))
        
        c.line(3 * cm, 4.8 * cm, 7 * cm, 4.8 * cm)
        c.line(12 * cm, 4.8 * cm, 17 * cm, 4.8 * cm)

        c.showPage()
        c.save()

        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=Faktura_{transaction_id}_{transaction["timestamp"]}.pdf'
        return response

    except Exception as e:
        logging.error(f"PDF yaratma xətası (ReportLab): {e}")
        flash(f'Faktura yaratmaq mümkün olmadı. Xəta: {e}', 'danger')
        return redirect(url_for('invoices'))
    finally:
        conn.close()

@app.route('/finance/invoices')
@login_required
@role_required('Admin')
def invoices():
    conn = get_db_connection()
    try:
        transactions = conn.execute('''
            SELECT 
                ft.*, 
                e.username AS employee_name -- İşçinin adını çəkirik
            FROM 
                financial_transactions ft 
            LEFT JOIN 
                employees e ON ft.employee_id = e.id
            ORDER BY 
                ft.timestamp DESC
        ''').fetchall()
        
        return render_template('invoices.html', transactions=transactions) 
    except Exception as e:
        app.logger.error(f'Fakturalar yüklənərkən xəta baş verdi: {str(e)}')
        flash('Fakturalar yüklənərkən xəta baş verdi.', 'danger')
        return redirect(url_for('index'))
    finally:
        try:
            conn.close()
        except:
            pass

@app.route('/finance/debts', methods=['GET', 'POST'])
@app.route('/finance/debts/<company_name>', methods=['GET'])
@login_required
@role_required('Admin')
def debts_management(company_name=None):
    """
    Borcların və Alacaqların İdarə Edilməsi (Debts and Receivables Management).
    :route /finance/debts - Əsas icmal (GET) və yeni qeyd/əməliyyatlar (POST)
    :route /finance/debts/<company_name> - Seçilmiş şirkətin detalları (GET)
    """
    conn = None
    try:
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        employee_id = session.get('user_id')
        current_time = get_current_baku_time()

        # --- POST Handling (Yeni Qeyd, Toplu Ödəniş, Əvəzləşdirmə) ---
        if request.method == 'POST':
            action = request.form.get('action')
            
            # ========== 1. YENİ QEYD ƏLAVƏ EDİLMƏSİ ==========
            if action == 'add':
                party_name = request.form.get('party_name').strip()
                record_type = request.form.get('type')  # 'debt' (Öhdəlik) və ya 'receivable' (Alacaq)
                initial_amount = request.form.get('initial_amount', type=float)
                due_date = request.form.get('due_date')

                if not all([party_name, record_type, initial_amount]) or initial_amount <= 0:
                    flash("Bütün sahələri doldurun və məbləğ 0-dan böyük olsun.", 'danger')
                    return redirect(url_for('debts_management'))

                # Məntiqi Təyinat: İstifadəçi tərəfindən daxil edilən növü DB-yə uyğunlaşdırma
                if record_type in ['debt']:
                    db_type = 'Debt'
                    display_type = "Öhdəlik (Borc)"
                elif record_type in ['receivable', 'credit']: # 'credit' də dəstəklənir
                    db_type = 'Receivable'
                    display_type = "Alacaq (Tələb)"
                else:
                    flash("Yanlış borc/alacaq növü.", 'danger')
                    return redirect(url_for('debts_management'))
                
                conn.execute('BEGIN')
                try:
                    conn.execute(
                        '''INSERT INTO debts (party_name, type, initial_amount, paid_amount, status, created_at, due_date) 
                           VALUES (?, ?, ?, 0, 'Pending', ?, ?)''',
                        (party_name, db_type, initial_amount, current_time, due_date if due_date else None)
                    )
                    log_msg = f"Yeni {display_type} qeydə alındı: '{party_name}', {initial_amount:.2f} AZN."
                    audit_log(conn, employee_id, log_msg)
                    conn.commit()
                    flash(f"'{party_name}' üçün {initial_amount:.2f} AZN məbləğində {display_type} uğurla qeydə alındı.", 'success')
                    return redirect(url_for('debts_management'))
                except sqlite3.Error as e:
                    conn.rollback()
                    flash(f"Qeydiyyat zamanı SQL xətası: {str(e)}", 'danger')
                    app.logger.error(f"Yeni Qeyd SQL xətası: {str(e)}", exc_info=True)
                    return redirect(url_for('debts_management'))


            # ========== 2. TOPLU ÖDƏNİŞ/QƏBUL ƏMƏLİYYATI ==========
            elif action == 'pay_consolidated':
                company_name = request.form.get('company_name').strip()
                payment_type = request.form.get('payment_type') # 'debt_payment' (Biz ödəyirik) və ya 'receivable_payment' (Biz qəbul edirik)
                payment_amount = request.form.get('payment_amount', type=float)
                payment_desc = request.form.get('payment_desc', 'Toplu ödəniş/qəbul əməliyyatı')

                if not company_name or payment_amount is None or payment_amount <= 0:
                    flash("Şirkət adını və 0-dan böyük məbləği daxil edin.", 'danger')
                    return redirect(url_for('debts_management', company_name=company_name, show_details='true'))
                
                records_to_pay = []
                transaction_type = None
                category = None
                log_prefix = ""
                
                conn.execute('BEGIN')
                try:
                    if payment_type == 'debt_payment': # Biz borc ödəyirik (Expense)
                        # Yalnız 'debts' cədvəlindən olan öhdəliklər
                        records_to_pay_query = '''
                            SELECT id, 'debts' as source_table, initial_amount, paid_amount, (initial_amount - paid_amount) as remaining, 'Debt' as type, created_at
                            FROM debts 
                            WHERE LOWER(TRIM(party_name)) = LOWER(?) 
                            AND type IN ('debt', 'Debt') 
                            AND (initial_amount - paid_amount) > 0.01 
                            ORDER BY created_at ASC
                        '''
                        records_to_pay = conn.execute(records_to_pay_query, (company_name.lower(),)).fetchall()
                        transaction_type = 'Expense'
                        category = 'Təchizatçı Borc Ödənişi'
                        log_prefix = "Borc Ödənişi"
                        
                    elif payment_type == 'receivable_payment': # Biz ödəniş qəbul edirik (Income)
                        # 'debts' cədvəlindəki alacaqlar
                        records_debts = conn.execute('''
                            SELECT id, 'debts' as source_table, initial_amount, paid_amount, (initial_amount - paid_amount) as remaining, 'Receivable' as type, created_at
                            FROM debts 
                            WHERE LOWER(TRIM(party_name)) = LOWER(?) 
                            AND type IN ('receivable', 'Receivable') 
                            AND (initial_amount - paid_amount) > 0.01 
                        ''', (company_name.lower(),)).fetchall()
                        
                        # 'customer_debts' cədvəlindəki alacaqlar
                        records_cust = conn.execute('''
                            SELECT id, 'customer_debts' as source_table, amount as initial_amount, paid_amount, (amount - paid_amount) as remaining, 'Receivable' as type, created_at
                            FROM customer_debts 
                            WHERE LOWER(TRIM(customer_name)) = LOWER(?) 
                            AND (amount - paid_amount) > 0.01 
                            AND status != 'Ödənilib' 
                        ''', (company_name.lower(),)).fetchall()

                        # Bütün alacaqları birləşdir və FIFO prinsipinə görə çeşidlə
                        records_to_pay = sorted([dict(r) for r in records_debts] + [dict(r) for r in records_cust], key=lambda x: x['created_at'])
                        
                        transaction_type = 'Income'
                        category = 'Müştəri Borc Qəbulu'
                        log_prefix = "Alacaq Qəbulu"
                    else:
                        conn.rollback()
                        flash("Yanlış ödəniş növü.", 'danger')
                        return redirect(url_for('debts_management', company_name=company_name, show_details='true'))
                    
                    
                    remaining_amount = payment_amount
                    total_logged = 0.0

                    for record in records_to_pay:
                        if remaining_amount <= 0.01: break
                        
                        pay_on_record = min(remaining_amount, record['remaining'])
                        new_paid_amount = record['paid_amount'] + pay_on_record
                        source_table = record['source_table'] 
                        
                        # Əsas cədvəli yenilə (Məntiq Düzgündür)
                        if source_table == 'debts':
                            conn.execute('UPDATE debts SET paid_amount = ? WHERE id = ?', (new_paid_amount, record['id']))
                            # float tolerantlığı ilə statusu yenilə
                            if new_paid_amount >= record['initial_amount'] - 0.01:
                                conn.execute("UPDATE debts SET status = 'Paid' WHERE id = ?", (record['id'],))
                        else: # customer_debts
                            conn.execute('UPDATE customer_debts SET paid_amount = ? WHERE id = ?', (new_paid_amount, record['id']))
                            # float tolerantlığı ilə statusu yenilə
                            if new_paid_amount >= record['initial_amount'] - 0.01:
                                conn.execute("UPDATE customer_debts SET status = 'Ödənilib' WHERE id = ?", (record['id'],))
                                
                        # Ödəniş tarixçəsi (Source table və transaction type ilə)
                        conn.execute(
                            '''INSERT INTO debt_payments 
                               (debt_id, amount, description, payment_date, employee_id, source_table, transaction_type) 
                               VALUES (?, ?, ?, ?, ?, ?, ?)''',
                            (record['id'], pay_on_record, payment_desc, current_time, employee_id, source_table, transaction_type)
                        )
                        
                        remaining_amount -= pay_on_record
                        total_logged += pay_on_record
                        
                    # Maliyyə Hərəkətlərinə qeyd
                    log_msg = f"{log_prefix} '{company_name}' tərəfindən/üçün: {total_logged:.2f} AZN. {payment_desc}"
                    
                    if total_logged > 0.01:
                        conn.execute(
                            '''INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id) 
                               VALUES (?, ?, ?, ?, ?, ?)''',
                            (transaction_type, category, total_logged, log_msg, current_time, employee_id)
                        )
                        audit_log(conn, employee_id, log_msg)
                        conn.commit()
                        flash(f"'{company_name}' tərəfi ilə {total_logged:.2f} AZN məbləğində əməliyyat uğurla qeyd edildi.", 'success')
                    else:
                        conn.rollback()
                        flash("Bu şirkətin ödəniləcək aktiv borc/alacaq qeydi tapılmadı.", 'info')

                    return redirect(url_for('debts_management', company_name=company_name, show_details='true'))

                except Exception as e:
                    conn.rollback()
                    flash(f"Ödəniş zamanı gözlənilməz xəta: {str(e)}", 'danger')
                    app.logger.error(f"Toplu Ödəniş xətası: {str(e)}", exc_info=True)
                    return redirect(url_for('debts_management', company_name=company_name, show_details='true'))

            
            # ========== 3. ƏVƏZLƏŞDİRMƏ ƏMƏLİYYATI ==========
            elif action == 'offset_debt_receivable':
                company_name = request.form.get('company_name').strip()
                offset_amount = request.form.get('offset_amount', type=float)
                offset_desc = request.form.get('offset_desc', 'Qarşılıqlı borc/alacaq əvəzləşdirilməsi')

                if not company_name or offset_amount is None or offset_amount <= 0:
                    flash("Şirkət adını və 0-dan böyük məbləği daxil edin.", 'danger')
                    return redirect(url_for('debts_management', company_name=company_name, show_details='true'))

                conn.execute('BEGIN')
                try:
                    # 1. Açıq Borc Qeydləri (Öhdəlik) - debts cədvəlindən
                    debt_records = conn.execute('''
                        SELECT id, initial_amount, paid_amount, (initial_amount - paid_amount) as remaining, 'debts' as source_table
                        FROM debts 
                        WHERE LOWER(TRIM(party_name)) = LOWER(?) 
                        AND type IN ('debt', 'Debt') 
                        AND (initial_amount - paid_amount) > 0.01 
                        ORDER BY created_at ASC
                    ''', (company_name.lower(),)).fetchall()
                    
                    # 2. Açıq Alacaq Qeydləri (Tələb) - debts və customer_debts cədvəllərindən
                    receivable_records_debts = conn.execute('''
                        SELECT id, 'debts' as source_table, initial_amount, paid_amount, (initial_amount - paid_amount) as remaining, created_at
                        FROM debts 
                        WHERE LOWER(TRIM(party_name)) = LOWER(?) 
                        AND type IN ('receivable', 'Receivable') 
                        AND (initial_amount - paid_amount) > 0.01 
                    ''', (company_name.lower(),)).fetchall()

                    receivable_records_customer_debts = conn.execute('''
                        SELECT id, 'customer_debts' as source_table, amount as initial_amount, paid_amount, (amount - paid_amount) as remaining, created_at
                        FROM customer_debts 
                        WHERE LOWER(TRIM(customer_name)) = LOWER(?) 
                        AND (amount - paid_amount) > 0.01 
                        AND status != 'Ödənilib'
                    ''', (company_name.lower(),)).fetchall()

                    all_receivable_records = [dict(r) for r in receivable_records_debts] + [dict(r) for r in receivable_records_customer_debts]
                    receivable_records = sorted(all_receivable_records, key=lambda x: x['created_at'])
                    
                    total_debt = sum(r['remaining'] for r in debt_records)
                    total_receivable = sum(r['remaining'] for r in receivable_records)
                    max_offset = min(total_debt, total_receivable)
                    
                    if offset_amount > max_offset + 0.01:
                        conn.rollback()
                        flash(f"Daxil edilən məbləğ ({offset_amount:.2f} AZN) maksimum əvəzləşdirmə məbləğindən ({max_offset:.2f} AZN) çoxdur.", 'danger')
                        return redirect(url_for('debts_management', company_name=company_name, show_details='true'))
                    
                    
                    # 3. Borcların (Öhdəlik) azaldılması
                    remaining_to_offset = offset_amount
                    for record in debt_records:
                        if remaining_to_offset <= 0.01: break
                        
                        pay_amount = min(remaining_to_offset, record['remaining'])
                        new_paid_amount = record['paid_amount'] + pay_amount
                        
                        conn.execute('UPDATE debts SET paid_amount = ? WHERE id = ?', (new_paid_amount, record['id']))
                        if new_paid_amount >= record['initial_amount'] - 0.01:
                            conn.execute("UPDATE debts SET status = 'Paid' WHERE id = ?", (record['id'],))
                        
                        conn.execute(
                            '''INSERT INTO debt_payments (debt_id, amount, description, payment_date, employee_id, source_table, transaction_type) 
                               VALUES (?, ?, ?, ?, ?, ?, ?)''',
                            (record['id'], pay_amount, offset_desc, current_time, employee_id, 'debts', 'Offset Expense')
                        )
                        remaining_to_offset -= pay_amount

                    # 4. Alacaqların (Tələb) azaldılması
                    remaining_to_offset = offset_amount # Məbləği yenidən sıfırlayırıq
                    for record in receivable_records:
                        if remaining_to_offset <= 0.01: break
                        
                        pay_amount = min(remaining_to_offset, record['remaining'])
                        new_paid_amount = record['paid_amount'] + pay_amount
                        source_table = record['source_table'] 
                        
                        if source_table == 'debts':
                            conn.execute('UPDATE debts SET paid_amount = ? WHERE id = ?', (new_paid_amount, record['id']))
                            if new_paid_amount >= record['initial_amount'] - 0.01:
                                conn.execute("UPDATE debts SET status = 'Paid' WHERE id = ?", (record['id'],))
                        else: # customer_debts
                            conn.execute('UPDATE customer_debts SET paid_amount = ? WHERE id = ?', (new_paid_amount, record['id']))
                            if new_paid_amount >= record['initial_amount'] - 0.01:
                                conn.execute("UPDATE customer_debts SET status = 'Ödənilib' WHERE id = ?", (record['id'],))
                        
                        conn.execute(
                            '''INSERT INTO debt_payments (debt_id, amount, description, payment_date, employee_id, source_table, transaction_type) 
                               VALUES (?, ?, ?, ?, ?, ?, ?)''',
                            (record['id'], pay_amount, offset_desc, current_time, employee_id, source_table, 'Offset Income')
                        )
                        remaining_to_offset -= pay_amount

                    # 5. Financial Transaction Log (Net sıfır təsir, Double-Entry)
                    log_msg = f"Qarşılıqlı əvəzləşdirmə '{company_name}' tərəfi ilə: {offset_amount:.2f} AZN. {offset_desc}"
                    
                    conn.execute(
                        '''INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id) VALUES (?, ?, ?, ?, ?, ?)''',
                        ('Expense', 'Əvəzləşdirmə (Borc Azalması)', offset_amount, log_msg, current_time, employee_id)
                    )
                    conn.execute(
                        '''INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id) VALUES (?, ?, ?, ?, ?, ?)''',
                        ('Income', 'Əvəzləşdirmə (Alacaq Azalması)', offset_amount, log_msg, current_time, employee_id)
                    )
                    
                    audit_log(conn, employee_id, log_msg) 
                    conn.commit()
                    flash(f"'{company_name}' tərəfi ilə {offset_amount:.2f} AZN məbləğində əvəzləşdirmə uğurla aparıldı.", 'success')
                    return redirect(url_for('debts_management', company_name=company_name, show_details='true'))

                except Exception as e:
                    conn.rollback()
                    app.logger.error(f"Offsetting Gözlənilməz xəta: {str(e)}", exc_info=True)
                    flash(f"Əvəzləşdirmə zamanı gözlənilməz xəta: {str(e)}", 'danger')
                    return redirect(url_for('debts_management', company_name=company_name, show_details='true'))


        # --- GET Handling (Əsas Siyahı/Detallı Görünüş) ---

        show_details = request.args.get('show_details') == 'true'

        if company_name and show_details:
            company_name_lower = company_name.strip().lower()

            # Detallı Görünüş: Seçilmiş Şirkətin Bütün Açıq Qeydləri
            # ******************* DÜZƏLİŞ EDİLMİŞ HİSSƏ *******************
            detailed_query = '''
                SELECT 
                    id, 'debts' as source_table, TRIM(party_name) as party_name, type, description,
                    initial_amount, paid_amount, (initial_amount - paid_amount) as remaining,
                    created_at, due_date
                FROM debts
                WHERE LOWER(TRIM(party_name)) = ?
                AND (initial_amount - paid_amount) > 0.01
                
                UNION ALL
                
                SELECT 
                    id, 'customer_debts' as source_table, TRIM(customer_name) as party_name, 'Receivable' as type, description,
                    amount as initial_amount, paid_amount, (amount - paid_amount) as remaining,
                    created_at, NULL as due_date  -- <<< due_date sütunu NULL olaraq əlavə edildi
                FROM customer_debts
                WHERE LOWER(TRIM(customer_name)) = ?
                AND (amount - paid_amount) > 0.01
                
                ORDER BY created_at ASC
            '''
            # *************************************************************
            
            # Parametrləri iki dəfə ötürürük, UNION ALL istifadə etdiyimiz üçün
            debt_details = conn.execute(detailed_query, (company_name_lower, company_name_lower)).fetchall()
            debt_details = [dict(d) for d in debt_details]
            
            # Ümumi Qalıq Balansların Hesablanması
            total_receivable_remaining = sum(d['remaining'] for d in debt_details if d['type'] == 'Receivable')
            total_debt_remaining = sum(d['remaining'] for d in debt_details if d['type'] == 'Debt')
            total_remaining = total_receivable_remaining - total_debt_remaining

            # Ödəniş Tarixçəsi Sorğusu
            payment_history_query = '''
                SELECT dp.*, e.username as employee_name, 
                       CASE 
                           WHEN dp.source_table = 'debts' THEN d.type 
                           WHEN dp.source_table = 'customer_debts' THEN 'Receivable'
                       END as debt_type_context
                FROM debt_payments dp
                LEFT JOIN employees e ON dp.employee_id = e.id
                LEFT JOIN debts d ON dp.source_table = 'debts' AND dp.debt_id = d.id
                LEFT JOIN customer_debts cd ON dp.source_table = 'customer_debts' AND dp.debt_id = cd.id
                WHERE (dp.source_table = 'debts' AND LOWER(TRIM(d.party_name)) = ?)
                OR (dp.source_table = 'customer_debts' AND LOWER(TRIM(cd.customer_name)) = ?)
                ORDER BY dp.payment_date DESC
            '''
            payment_history = conn.execute(payment_history_query, (company_name_lower, company_name_lower)).fetchall()
            payment_history = [dict(p) for p in payment_history]
            
            return render_template('debts_management.html', 
                                   selected_company=company_name,
                                   show_details=True,
                                   debt_details=debt_details,
                                   total_receivable_remaining=total_receivable_remaining,
                                   total_debt_remaining=total_debt_remaining,
                                   total_remaining=total_remaining,
                                   payment_history=payment_history
                                   )

        else:
            # Əsas Görünüş: Bütün Şirkətlər üzrə Toplu Məlumatlar
            
            company_list_query = '''
                -- Debts Table (Borclar və Alacaqlar)
                SELECT TRIM(party_name) as name, type, initial_amount, paid_amount 
                FROM debts 
                WHERE (initial_amount - paid_amount) > 0.01 
                
                UNION ALL
                
                -- Customer Debts Table (Yalnız Alacaqlar)
                SELECT TRIM(customer_name) as name, 'Receivable' as type, amount as initial_amount, paid_amount 
                FROM customer_debts 
                WHERE (amount - paid_amount) > 0.01 AND status != 'Ödənilib'
            '''
            all_records = conn.execute(company_list_query).fetchall()

            companies_data = {}
            for rec in all_records:
                company_name_key = rec['name'].strip()
                initial = rec['initial_amount']
                paid = rec['paid_amount']
                remaining = initial - paid
                
                if company_name_key not in companies_data:
                    companies_data[company_name_key] = {
                        'company_name': company_name_key,
                        'total_receivable': 0.0,
                        'total_debt': 0.0,
                        'debt_count': 0
                    }
                
                if rec['type'] in ['receivable', 'Receivable']:
                    companies_data[company_name_key]['total_receivable'] += remaining
                else: # 'debt', 'Debt'
                    companies_data[company_name_key]['total_debt'] += remaining
                    
                companies_data[company_name_key]['debt_count'] += 1

            companies = list(companies_data.values())
            
            # Xalis balansa görə çeşidləmə (Ən böyük borcdan ən böyük alacağa)
            companies.sort(key=lambda x: (x['total_receivable'] - x['total_debt']))

            return render_template('debts_management.html', 
                                   companies=companies,
                                   show_details=False)

    except Exception as e:
        if 'conn' in locals() and conn:
            try:
                # Tranzaksiya halında rollback, əlaqə açıqdırsa bağla
                conn.rollback()
            except:
                pass
        app.logger.error(f'Borcların idarə edilməsi səhifəsi yüklənərkən xəta: {str(e)}', exc_info=True)
        flash('Səhifə yüklənərkən gözlənilməz xəta baş verdi.', 'danger')
        return redirect(url_for('index'))
    finally:
        if 'conn' in locals() and conn:
            # Əlaqəni bağla
            try:
                conn.close()
            except:
                pass


@app.route('/finance/pay-debt/<party_name>', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def pay_consolidated_debt(party_name):
    """
    Seçilmiş tərəfin (party_name) fərdi borc/alacaq qeydləri üzrə ödənişin qeydiyyatı.
    
    POST: Fərdi qeyd üzrə ödənişi icra edir, əsas cədvəli, ödəniş tarixçəsini və 
          maliyyə hərəkətlərini yeniləyir.
    GET: Həmin tərəfin bütün açıq qalıq borclarını/alacaqlarını siyahılayır.
    """
    conn = None
    try:
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        employee_id = session.get('user_id')
        current_time = get_current_baku_time()

        if not employee_id:
            flash('İstifadəçi məlumatları tapılmadı. Zəhmət olmasa yenidən daxil olun.', 'danger')
            return redirect(url_for('login'))
        
        # Məntiqi yoxlamalar üçün party_name-i təmizlə
        cleaned_party_name = party_name.strip()

        if request.method == 'POST':
            record_id = request.form.get('record_id', type=int)
            source_table = request.form.get('source_table') # debts və ya customer_debts
            payment_amount = request.form.get('payment_amount', type=float)
            payment_desc = request.form.get('payment_desc', 'Fərdi qeyd üzrə ödəniş')

            if not all([record_id, source_table]) or payment_amount is None or payment_amount <= 0:
                flash("Ödəniş üçün qeyd ID-si, cədvəl mənbəyi və ya düzgün məbləğ daxil edin.", 'danger')
                return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))
            
            initial_amount = 0.0
            current_paid = 0.0
            debt_type = None
            
            # 1. Qeydin tapılması və qalıq məbləğin hesablanması
            if source_table == 'debts':
                # Burada 'party_name' yoxlamasını da edirik
                record_query = 'SELECT initial_amount, paid_amount, type FROM debts WHERE id = ? AND LOWER(TRIM(party_name)) = LOWER(TRIM(?))'
                record = conn.execute(record_query, (record_id, cleaned_party_name)).fetchone()
                
                if not record:
                    flash(f"ID {record_id} olan maliyyə qeydi ('debts') tapılmadı və ya '{cleaned_party_name}' tərəfinə aid deyil.", 'danger')
                    return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))
                    
                current_paid = record['paid_amount']
                initial_amount = record['initial_amount']
                debt_type = record['type'] # 'Debt' (Expense) və ya 'Receivable' (Income)
                
            elif source_table == 'customer_debts':
                # Burada 'customer_name' yoxlamasını da edirik
                record_query = 'SELECT amount, paid_amount FROM customer_debts WHERE id = ? AND LOWER(TRIM(customer_name)) = LOWER(TRIM(?))'
                record = conn.execute(record_query, (record_id, cleaned_party_name)).fetchone()

                if not record:
                    flash(f"ID {record_id} olan maliyyə qeydi ('customer_debts') tapılmadı və ya '{cleaned_party_name}' tərəfinə aid deyil.", 'danger')
                    return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))

                current_paid = record['paid_amount']
                initial_amount = record['amount']
                debt_type = 'Receivable' # customer_debts həmişə alacaqdır
            
            else:
                flash("Yanlış cədvəl mənbəyi göstərilib.", 'danger')
                return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))
            
            
            remaining = initial_amount - current_paid
            new_paid_amount = current_paid + payment_amount
            
            if payment_amount > remaining + 0.01: # Float xətası toleransı
                flash(f"Daxil edilən məbləğ ({payment_amount:.2f} AZN) qalıqdan ({remaining:.2f} AZN) çoxdur.", 'danger')
                return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))
            
            # 2. Ödənişin qeydiyyatı (Transaction)
            conn.execute('BEGIN')
            try:
                log_msg = ""
                transaction_type = ""
                category = ""
                
                # A. Əsas cədvəlin yenilənməsi
                if source_table == 'debts':
                    conn.execute(
                        'UPDATE debts SET paid_amount = ? WHERE id = ?',
                        (new_paid_amount, record_id)
                    )
                    if new_paid_amount >= initial_amount - 0.01:
                        conn.execute("UPDATE debts SET status = 'Paid' WHERE id = ?", (record_id,))

                    if debt_type == 'Receivable':
                        transaction_type = 'Income' # Bizə pul gəldi
                        category = 'Müştəri Alacağı Ödənişi'
                        log_msg = f"Müştəri alacağı ödənişi (Debts): {cleaned_party_name} üçün {record_id} ID - {payment_amount:.2f} AZN"
                    else: # 'Debt'
                        transaction_type = 'Expense' # Bizdən pul çıxdı
                        category = 'Təchizatçı Borc Ödənişi'
                        log_msg = f"Təchizatçı borcu ödənişi (Debts): {cleaned_party_name} üçün {record_id} ID - {payment_amount:.2f} AZN"
                        
                elif source_table == 'customer_debts':
                    conn.execute(
                        'UPDATE customer_debts SET paid_amount = ? WHERE id = ?',
                        (new_paid_amount, record_id)
                    )
                    if new_paid_amount >= initial_amount - 0.01:
                        conn.execute("UPDATE customer_debts SET status = 'Ödənilib' WHERE id = ?", (record_id,))

                    transaction_type = 'Income' # Bizə pul gəldi
                    category = 'Müştəri Alacağı Ödənişi'
                    log_msg = f"Müştəri alacağı ödənişi (Customer Debts): {cleaned_party_name} üçün {record_id} ID - {payment_amount:.2f} AZN"

                # B. Maliyyə Hərəkətlərinə qeyd
                conn.execute(
                    '''INSERT INTO financial_transactions (type, category, amount, description, timestamp, employee_id) 
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (transaction_type, category, payment_amount, log_msg + f" (Qeyd: {payment_desc})", current_time, employee_id)
                )
                
                # C. Ödəniş Tarixçəsinə qeyd
                # *** DİQQƏT: Buraya 'transaction_type' sahəsini əlavə edirəm. ***
                conn.execute(
                    '''INSERT INTO debt_payments 
                       (debt_id, amount, description, payment_date, employee_id, source_table, transaction_type) 
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (record_id, payment_amount, payment_desc, current_time, employee_id, source_table, transaction_type)
                )
                
                # D. Audit
                audit_log(conn, employee_id, log_msg)
                
                conn.commit()
                flash(f"'{cleaned_party_name}' şirkəti üçün {payment_amount:.2f} AZN məbləğində ödəniş uğurla qeyd edildi.", 'success')
                return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))

            except sqlite3.Error as e:
                conn.rollback()
                app.logger.error(f"Ödəniş qeyd edilərkən SQL xətası: {str(e)}", exc_info=True)
                flash(f"Ödəniş qeyd edilərkən SQL xətası baş verdi: {str(e)}", 'danger')
                return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Ödəniş qeyd edilərkən gözlənilməz xəta: {str(e)}", exc_info=True)
                flash(f"Ödəniş qeyd edilərkən gözlənilməz xəta: {str(e)}", 'danger')
                return redirect(url_for('pay_consolidated_debt', party_name=cleaned_party_name))

        # GET Request: Bütün qalıq borcları/alacaqları siyahıya almaq
        records_query = '''
            -- 1. Records from 'debts' table (Debt/Receivable)
            SELECT 
                id,
                'debts' as source_table,
                TRIM(party_name) as party_name,
                type,
                description,
                initial_amount,
                paid_amount,
                (initial_amount - paid_amount) as remaining,
                created_at
            FROM debts
            WHERE LOWER(TRIM(party_name)) = LOWER(TRIM(?))
            AND (initial_amount - paid_amount) > 0.01

            UNION ALL

            -- 2. Records from 'customer_debts' table (Receivable only)
            SELECT 
                id,
                'customer_debts' as source_table,
                TRIM(customer_name) as party_name,
                'Receivable' as type,
                description,
                amount as initial_amount, -- Use 'amount' for initial display
                paid_amount,
                (amount - paid_amount) as remaining,
                created_at
            FROM customer_debts
            WHERE LOWER(TRIM(customer_name)) = LOWER(TRIM(?))
            AND status != 'Ödənilib'
            AND (amount - paid_amount) > 0.01

            ORDER BY 
                CASE 
                    WHEN type IN ('receivable', 'Receivable') THEN 0 -- Alacaqlar (Müştəri ödəyir) öncə gəlsin
                    WHEN type IN ('debt', 'Debt') THEN 1              -- Borclar (Biz ödəyirik) sonra gəlsin
                    ELSE 2
                END,
                created_at ASC
        '''
        # party_name dəyişəni iki dəfə ötürülür (UNION ALL olduğu üçün)
        records = conn.execute(records_query, (cleaned_party_name, cleaned_party_name)).fetchall()
        
        records = [dict(rec) for rec in records]

        return render_template('pay_debt.html',
                               party_name=cleaned_party_name,
                               records=records,
                               has_balance=len(records) > 0)

    except Exception as e:
        if 'conn' in locals() and conn:
            conn.rollback()
        app.logger.error(f'Ödəniş səhifəsi yüklənərkən xəta: {str(e)}', exc_info=True)
        flash('Ödəniş səhifəsi yüklənərkən xəta baş verdi. Zəhmət olmasa yenidən cəhd edin.', 'danger')
        return redirect(url_for('debts_management')) 
    finally:
        if 'conn' in locals() and conn:
            try:
                conn.close()
            except:
                pass

from io import StringIO
@app.route('/production', methods=['GET'])
@login_required
def production_form():
    """
    İstehsalat əməliyyatı formasını göstərir və bütün aktiv məhsulların siyahısını yükləyir.
    """
    conn = get_db_connection()
    products_for_json = []
    try:
        products_list = conn.execute("""
            SELECT id, name, unit, stock, price 
            FROM products 
            WHERE is_active = 1
            ORDER BY name
        """).fetchall()
        
        # Siyahını Jinja2/JS üçün uyğun formatda qaytarır
        products_for_json = [dict(p) for p in products_list]
        
    except Exception as e:
        app.logger.error(f"Məhsul siyahısını çəkərkən xəta: {e}", exc_info=True)
        flash('Məhsul məlumatları yüklənərkən xəta baş verdi.', 'danger')
        
    finally:
        conn.close()
        
    return render_template('production.html', products=products_for_json)

def validate_production_data(conn, input_items, output_items):
    """Validate production data before processing."""
    errors = []
    
    # Check input items
    for item in input_items:
        product = conn.execute("SELECT id, name, stock FROM products WHERE id = ?", 
                            (item['id'],)).fetchone()
        if not product:
            errors.append(f"Xəta: ID {item['id']} ilə xammal məhsul tapılmadı")
            continue
            
        if product['stock'] < item['quantity']:
            errors.append(
                f"Kifayət qədər stok yoxdur: {product['name']}. "
                f"Mövcud: {product['stock']}, Tələb olunan: {item['quantity']}"
            )
    
    # Check output items
    for item in output_items:
        # For existing products, check if they exist and have valid data
        if not item.get('is_new'):
            product_id = item.get('product_id') or item.get('id')
            if not product_id:
                errors.append("Mövcud məhsul üçün məhsul ID-si daxil edilməyib")
                continue
                
            product = conn.execute("SELECT id, name FROM products WHERE id = ?", 
                                (product_id,)).fetchone()
            if not product:
                errors.append(f"ID {product_id} ilə məhsul tapılmadı")
        # For new products, ensure we have all required fields
        elif item.get('is_new') and not item.get('new_name'):
            errors.append("Yeni məhsul üçün ad daxil edilməyib")
        elif not item.get('is_new') and not (item.get('product_id') or item.get('id')):
            errors.append("Mövcud məhsul üçün keçərli məhsul seçilməyib")
    
    return errors
# Köməkçi funksiya: Listdəki elementi etibarlı şəkildə almaq
def safe_get(lst, index, default=''):
    """Listin verilmiş indeksindəki dəyəri qaytarır, yoxsa boş sətir və ya default dəyər."""
    # Sadəlik naminə, bu funksiya kodun daxilində deyil,
    # yuxarıda və ya bir utility faylında təyin olunmalıdır.
    return lst[index].strip() if index < len(lst) else default


@app.route('/production', methods=['POST'])
@login_required
def handle_production():
    app.logger.info("=== Yeni istehsal əməliyyatı başladı ===")
    
    employee_id = session.get('user_id')
    app.logger.info(f"İşçi ID: {employee_id}")
    
    # Form məlumatlarını al
    input_product_ids_str = request.form.getlist('input_product_id[]')
    input_quantities_str = request.form.getlist('input_quantity[]')
    
    output_product_ids_str = request.form.getlist('output_product_id[]')
    output_new_product_names = request.form.getlist('output_new_product_name[]')
    output_units = request.form.getlist('output_unit[]')
    output_quantities_str = request.form.getlist('output_quantity[]')
    
    description = request.form.get('description', '').strip()
    
    app.logger.info(f"Giriş məhsulları (ID): {input_product_ids_str}")
    app.logger.info(f"Giriş miqdarı: {input_quantities_str}")
    app.logger.info(f"Çıxış məhsul ID-ləri: {output_product_ids_str}")
    app.logger.info(f"Yeni məhsul adları: {output_new_product_names}")
    app.logger.info(f"Çıxış vahidləri: {output_units}")
    app.logger.info(f"Çıxış miqdarı: {output_quantities_str}")
    
    if not employee_id:
        flash("Sessiya başa çatıb. Zəhmət olmasa yenidən daxil olun.", 'danger')
        return redirect(url_for('login'))
        
    conn = get_db_connection()
    current_time = get_current_baku_time()
    
    try:
        conn.execute('BEGIN TRANSACTION') 

        # =============================================================
        # 1. INPUT (GİRİŞ) EMALI
        # =============================================================

        input_items = []
        seen_products = set()
        
        for i in range(max(len(input_product_ids_str), len(input_quantities_str))):
            try:
                # safe_get funksiyasını burada istifadə edirik
                id_str = safe_get(input_product_ids_str, i)
                qty_str = safe_get(input_quantities_str, i)
                
                if not id_str or not qty_str:
                    continue
                    
                product_id = int(id_str)
                quantity = float(qty_str.replace(',', '.'))

                if quantity <= 0:
                    continue

                # Eyni xammal iki dəfə seçilibsə → miqdarı topla
                if product_id in seen_products:
                    existing = next((x for x in input_items if x['id'] == product_id), None)
                    if existing:
                        existing['quantity'] += quantity
                        continue

                input_items.append({
                    'id': product_id,
                    'quantity': quantity
                })
                seen_products.add(product_id)
                
            except Exception as e:
                app.logger.error(f'Input emal xətası: {e}', exc_info=True)
                continue

        # =============================================================
        # 2. OUTPUT (ÇIXIŞ) — TƏKMİLLƏŞDİRİLMİŞ EMAL
        # =============================================================

        output_items = []
        max_len = max(len(output_product_ids_str), 
                      len(output_new_product_names), 
                      len(output_units), 
                      len(output_quantities_str))
        
        for i in range(max_len):
            
            # safe_get ilə bütün listləri maksimum uzunluqla götürürük
            product_id_str = safe_get(output_product_ids_str, i)
            new_name = safe_get(output_new_product_names, i)
            unit = safe_get(output_units, i, default='ədəd')
            qty_str = safe_get(output_quantities_str, i)

            if not qty_str:
                continue
                
            try:
                quantity = float(qty_str.replace(',', '.'))
                if quantity <= 0:
                    continue
            except:
                continue
            
            # 1. YENİ MƏHSUL MƏNTİQİ: product_id rəqəm deyilsə (məsələn 'new') VƏ yeni ad verilibsə
            if not product_id_str.isdigit() and new_name:
                output_items.append({
                    'is_new': True,
                    'new_name': new_name,
                    'quantity': quantity,
                    'unit': unit
                })
            # 2. MÖVCUD MƏHSUL MƏNTİQİ: product_id rəqəmdirsə
            elif product_id_str.isdigit():
                output_items.append({
                    'is_new': False,
                    'product_id': int(product_id_str),
                    'quantity': quantity,
                    'unit': unit 
                })
            # 3. Ziddiyyətli və ya boş sətirlər atlanır
            else:
                continue

        app.logger.info(f"FINAL OUTPUT ITEMS (DÜZƏLDİLMİŞ): {output_items}")

        # =============================================================
        # 3.1. Eyni məhsulları birləşdir (Məcburi addım, çünki formda eyni məhsul bir neçə dəfə əlavə oluna bilər)
        # =============================================================
        consolidated_output = {}
        for item in output_items:
            if item.get("is_new"):
                new_key = f"new_{item.get('new_name', '')}_{item.get('unit', '')}"
                if new_key in consolidated_output:
                    consolidated_output[new_key]['quantity'] += item['quantity']
                else:
                    consolidated_output[new_key] = item
            else:
                product_id = item['product_id']
                if product_id in consolidated_output:
                    consolidated_output[product_id]['quantity'] += item['quantity']
                else:
                    consolidated_output[product_id] = item

        output_items = list(consolidated_output.values())
        app.logger.info(f"BİRLƏŞDİRİLMİŞ ÇIXIŞ MƏHSULLARI: {output_items}")

        # =============================================================
        # 3.2. Validasiya
        # =============================================================

        if not input_items:
            conn.rollback()
            flash('Ən azı bir xammal daxil edilməlidir!', 'danger')
            return redirect(url_for('production_form'))

        if not output_items:
            conn.rollback()
            flash('Ən azı bir hazır məhsul daxil edilməlidir!', 'danger')
            return redirect(url_for('production_form'))

        validation_errors = validate_production_data(conn, input_items, output_items)
        if validation_errors:
            conn.rollback()
            for err in validation_errors:
                flash(err, 'danger')
            return redirect(url_for('production_form'))

        # =============================================================
        # 4. Transfer Yarat
        # =============================================================
        
        cursor = conn.execute("""
            INSERT INTO production_transfers 
            (user_id, description, transfer_date, created_at)
            VALUES (?, ?, ?, ?)
        """, (employee_id, description, current_time, current_time))

        transfer_id = cursor.lastrowid
        if not transfer_id:
            raise Exception("Transfer qeydi yaradıla bilmədi")

        # =============================================================
        # 5. INPUT STOCK UPDATE (Xammal İstehlakı - Azalma)
        # =============================================================

        input_items_log = []

        for item in input_items:
            product_id = item['id']
            quantity = item['quantity']

            product = conn.execute("""
                SELECT id, name, unit, stock, price FROM products WHERE id = ?
            """, (product_id,)).fetchone()

            if not product:
                raise ValueError(f"ID {product_id} xammal tapılmadı")

            if product['stock'] < quantity:
                raise ValueError(
                    f'"{product["name"]}" üçün kifayət qədər stok yoxdur. '
                    f'Mövcud: {product["stock"]}, Tələb: {quantity}'
                )

            conn.execute("""
                UPDATE products SET stock = stock - ?, updated_at = ?
                WHERE id = ?
            """, (quantity, current_time, product_id))

            conn.execute("""
                INSERT INTO production_transfer_items
                (transfer_id, product_id, quantity, type)
                VALUES (?, ?, ?, 'Input')
            """, (transfer_id, product_id, quantity))

            record_transaction(
                conn, product_id, 'production_out',
                quantity, product['price'], product['price'] * quantity,
                employee_id
            )

            input_items_log.append(f"{product['name']} (-{quantity} {product['unit']})")

        # =============================================================
        # 6. OUTPUT EMALI (Hazır Məhsul Daxilolması - Artma)
        # =============================================================

        output_items_log = []

        for item in output_items:
            quantity = item['quantity']

            # Yeni məhsul
            if item.get("is_new"):
                new_name = item["new_name"]
                unit = item["unit"]

                existing = conn.execute(
                    "SELECT id FROM products WHERE name = ?", (new_name,)
                ).fetchone()

                if existing:
                    raise ValueError(f"'{new_name}' adlı məhsul artıq mövcuddur!")

                sku = generate_unique_sku("PRD")

                cursor = conn.execute("""
                    INSERT INTO products
                    (name, sku, category, unit, price, stock, min_stock, location, description, is_active, created_at, updated_at)
                    VALUES (?, ?, 'İstehsal', ?, 0.0, ?, 0, 'İstehsal anbarı', 'İstehsalda hazırlanıb', 1, ?, ?)
                """, (new_name, sku, unit, quantity, current_time, current_time))

                final_product_id = cursor.lastrowid
                product_info = conn.execute(
                    "SELECT id, name, unit FROM products WHERE id = ?", (final_product_id,)
                ).fetchone()
                
                if not product_info:
                    raise Exception(f"Yeni məhsul ({new_name}) yaradıla bilmədi.")
                
                final_product_id = product_info['id']
                unit = product_info['unit']

                conn.execute("""
                    INSERT INTO production_transfer_items
                    (transfer_id, product_id, quantity, type)
                    VALUES (?, ?, ?, 'Output')
                """, (transfer_id, final_product_id, quantity))

                record_transaction(
                    conn, final_product_id, 'production_in',
                    quantity, 0.0, 0.0, employee_id,
                    note=f"Yeni məhsul istehsalı: {new_name}"
                )

                output_items_log.append(f"{new_name} (+{quantity} {unit})")
                continue

            # Mövcud məhsul
            product_id = item['product_id']

            product = conn.execute("""
                SELECT id, name, unit, price FROM products WHERE id = ?
            """, (product_id,)).fetchone()

            if not product:
                raise ValueError(f"ID {product_id} məhsul tapılmadı")

            # Sadə toplama əməliyyatı
            conn.execute("""
                UPDATE products SET stock = stock + ?, updated_at = ?
                WHERE id = ?
            """, (quantity, current_time, product_id))

            conn.execute("""
                INSERT INTO production_transfer_items
                (transfer_id, product_id, quantity, type)
                VALUES (?, ?, ?, 'Output')
            """, (transfer_id, product_id, quantity))

            record_transaction(
                conn, product_id, 'production_in',
                quantity, product['price'], quantity * product['price'],
                employee_id,
                note=f"İstehsal daxilolması: {product['name']}"
            )

            output_items_log.append(f"{product['name']} (+{quantity} {product['unit']})")

        # =============================================================
        # 7. ƏMƏLİYYATI YEKUNLAŞDIR
        # =============================================================

        log_message = f"İstehsalat əməliyyatı #{transfer_id}: Xammal: {', '.join(input_items_log)}. Hazır məhsullar: {', '.join(output_items_log)}"
        audit_log(conn, employee_id, log_message)
        conn.commit()

        flash("İstehsalat əməliyyatı uğurla tamamlandı!", "success")
        return redirect(url_for("production_form"))

    except Exception as e:
        conn.rollback()
        app.logger.error(f'Ümumi xəta: {e}', exc_info=True)
        flash(str(e), 'danger')
        return redirect(url_for('production_form'))

    finally:
        conn.close()
from collections import defaultdict
import math
from jinja2.ext import do 
@app.route('/production/report', methods=['GET'])
@login_required 
def production_report():
    
    PER_PAGE = 50 

    # DƏYİŞİKLİK 1: Yeni Axtarış və Filter Parametrlərini Qəbul Et
    search_query = request.args.get('search_query', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    input_product_name = request.args.get('input_product_name', '').strip()
    output_product_name = request.args.get('output_product_name', '').strip()
    
    # Səhifə nömrəsini təhlükəsiz şəkildə integerə çevirmək
    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
        
    if page < 1:
        page = 1

    offset = (page - 1) * PER_PAGE
    
    conn = get_db_connection()
    production_transfers = []
    item_details = defaultdict(lambda: {'Input': [], 'Output': []})
    total_records = 0
    total_pages = 0
    
    try:
        # DƏYİŞİKLİK 2: WHERE hissəsini və parametrləri dinamik qur
        where_conditions = []
        params = []
        
        # 1. Ümumi Axtarış
        if search_query:
            search_term = f"%{search_query}%"
            where_conditions.append(
                """(
                    CAST(pt.id AS TEXT) LIKE ? OR
                    pt.description LIKE ? OR
                    u.username LIKE ?
                )"""
            )
            params.extend([search_term, search_term, search_term])
        
        # 2. Tarix Filteri
        if start_date:
            where_conditions.append("pt.transfer_date >= ?")
            params.append(start_date)
            
        if end_date:
            # End_date üçün axtarışı həmin günün sonuna qədər etmək üçün ' 23:59:59' əlavə etmək daha yaxşı olar
            # Lakin SQLitenin yalniz tarix (YYYY-MM-DD) formatı ilə müqayisəni nəzərə alaraq, sadəcə <= istifadə edirik
            # əgər transfer_date timestamp olaraq saxlanilirsa, bu düzgün olmaya bilər. 
            # Düzgün olması üçün 'YYYY-MM-DD 23:59:59' formatında tam timestamp tələb olunur.
            # Sadəlik üçün YYYY-MM-DD ilə müqayisəni saxlayırıq.
            where_conditions.append("pt.transfer_date <= ?")
            params.append(end_date)
            
        # DƏYİŞİKLİK 3: Xammal və Məhsul Adına görə Filtrləmə (Alt-Sorğularla)
        # Bu, daha mürəkkəbdir, çünki eyni transferdə həm Xammal, həm də Məhsul ola bilər.
        
        # 3. Xammal (Input) Adına görə Filter
        if input_product_name:
            search_term = f"%{input_product_name}%"
            where_conditions.append("""
                pt.id IN (
                    SELECT pti.transfer_id 
                    FROM production_transfer_items pti
                    JOIN products p ON pti.product_id = p.id
                    WHERE pti.type = 'Input' AND p.name LIKE ?
                )
            """)
            params.append(search_term)

        # 4. Hazır Məhsul (Output) Adına görə Filter
        if output_product_name:
            search_term = f"%{output_product_name}%"
            where_conditions.append("""
                pt.id IN (
                    SELECT pti.transfer_id 
                    FROM production_transfer_items pti
                    JOIN products p ON pti.product_id = p.id
                    WHERE pti.type = 'Output' AND p.name LIKE ?
                )
            """)
            params.append(search_term)
            
        # WHERE hissəsini birləşdir
        where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""

        # 1. Cəmi Qeyd Sayını Hesabla
        count_data = conn.execute(f"""
            SELECT COUNT(DISTINCT pt.id)
            FROM production_transfers pt
            JOIN employees u ON pt.user_id = u.id
            {where_clause}
        """, params).fetchone()
        
        total_records = count_data[0]

        # Səhifələmə məntiqi (əvvəlki kimi)
        total_pages = math.ceil(total_records / PER_PAGE) if total_records > 0 else 0
        
        if total_pages > 0 and page > total_pages:
            page = total_pages
            offset = (page - 1) * PER_PAGE
        elif total_pages == 0:
            page = 1
            offset = 0

        # 2. Əsas Köçürmələr Sorğusu (LIMIT və OFFSET ilə)
        # Burada artıq JOIN etməyə ehtiyac yoxdur, çünki axtarış alt-sorğularla həll olunub.
        transfers_data = conn.execute(f"""
            SELECT 
                pt.id AS transfer_id,
                pt.description,
                pt.transfer_date,
                u.username AS employee_name
            FROM production_transfers pt
            JOIN employees u ON pt.user_id = u.id
            {where_clause}
            ORDER BY pt.transfer_date DESC
            LIMIT ? OFFSET ?
        """, params + [PER_PAGE, offset]).fetchall()

        # ... (flash mesajları və məlumat boşdursa xəbərdarlıq qismi əvvəlki kimi qalır)
        
        # 3. Əməliyyat Detallarını Çək (Əvvəlki kimi qalır)
        if transfers_data:
            transfer_ids = tuple(t["transfer_id"] for t in transfers_data)
            placeholders = ",".join("?" for _ in transfer_ids)

            items_data = conn.execute(f"""
                SELECT 
                    pti.transfer_id,
                    pti.quantity,
                    pti.type,
                    COALESCE(p.name, 'Silinmiş Məhsul') AS product_name,
                    COALESCE(p.unit, 'ədəd') AS unit
                FROM production_transfer_items pti
                LEFT JOIN products p ON pti.product_id = p.id
                WHERE pti.transfer_id IN ({placeholders})
                ORDER BY pti.transfer_id, pti.type, COALESCE(p.name, '') DESC
            """, transfer_ids).fetchall()

            for item in items_data:
                item_details[item["transfer_id"]][item["type"]].append(item)

            production_transfers = [dict(t) for t in transfers_data]

    except sqlite3.Error as e:
        app.logger.error(f"İstehsalat hesabatı məlumatlarını çəkərkən xəta: {e}", exc_info=True)
        flash('Hesabat məlumatları yüklənərkən verilənlər bazası xətası baş verdi.', 'danger')
        
    finally:
        if conn:
            conn.close()

    # Pagination obyekti
    pagination = {
        # ... (əvvəlki kimi)
        'current_page': page,
        'total_pages': total_pages,
        'total_items': total_records,
        'has_prev': page > 1,
        'prev_num': page - 1,
        'has_next': page < total_pages,
        'next_num': page + 1,
        'per_page': PER_PAGE
    }

    # DƏYİŞİKLİK 4: Yeni filter dəyərlərini şablona göndər
    return render_template(
        'production_report.html',
        transfers=production_transfers,
        item_details=item_details,
        pagination=pagination,
        search_query=search_query,
        start_date=start_date,
        end_date=end_date,
        input_product_name=input_product_name,
        output_product_name=output_product_name
    )
@app.route('/production/export_xlsx', methods=['GET'])
@login_required
def export_production_to_excel():
    
    from io import BytesIO 

    conn = get_db_connection()
    try:
        transfers_query = """
            SELECT 
                pt.id AS "Transfer ID", 
                pt.description AS "Qeyd", 
                pt.transfer_date AS "Tarix", 
                u.username AS "Əməliyyatçı"
            FROM production_transfers pt
            JOIN employees u ON pt.user_id = u.id 
            ORDER BY pt.transfer_date DESC
        """
        transfers_data = conn.execute(transfers_query).fetchall()

        if not transfers_data:
            flash('İxrac ediləcək heç bir istehsalat əməliyyatı tapılmadı.', 'warning')
            return redirect(url_for('production_report'))

        transfer_ids = tuple(t['Transfer ID'] for t in transfers_data)
        placeholders = ','.join('?' for _ in transfer_ids)
        
        transfer_cols = ["Transfer ID", "Qeyd", "Tarix", "Əməliyyatçı"]
        transfers_df = pd.DataFrame(transfers_data, columns=transfer_cols)
        

        items_data = conn.execute(f"""
            SELECT 
                pti.transfer_id, 
                p.name AS product_name, 
                pti.quantity, 
                p.unit, 
                pti.type
            FROM production_transfer_items pti
            JOIN products p ON pti.product_id = p.id
            WHERE pti.transfer_id IN ({placeholders})
            ORDER BY pti.transfer_id DESC, pti.type DESC
        """, transfer_ids).fetchall()

        items_cols = ['transfer_id', 'product_name', 'quantity', 'unit', 'type']
        items_df = pd.DataFrame(items_data, columns=items_cols)
        
        items_df['Məhsul Detalı'] = (
            items_df['product_name'] + ': ' + 
            items_df['quantity'].round(3).astype(str) + ' ' + 
            items_df['unit']
        )
        
        

        input_df = items_df[items_df['type'] == 'Input']
        input_grouped = input_df.groupby('transfer_id')['Məhsul Detalı'].apply(
            lambda x: ' | '.join(x) 
        ).reset_index(name='Xammal (İstifadə Olunan)')
        input_grouped.rename(columns={'transfer_id': 'Transfer ID'}, inplace=True)
        
        output_df = items_df[items_df['type'] == 'Output']
        output_grouped = output_df.groupby('transfer_id')['Məhsul Detalı'].apply(
            lambda x: ' | '.join(x)
        ).reset_index(name='Məhsul (İstehsal Olunan)')
        output_grouped.rename(columns={'transfer_id': 'Transfer ID'}, inplace=True)

        
        final_df = transfers_df.merge(input_grouped, on='Transfer ID', how='left')
        
        final_df = final_df.merge(output_grouped, on='Transfer ID', how='left')
        
        final_df = final_df.fillna('—')
        
        output = BytesIO() 
        excel_buffer = pd.ExcelWriter(output, engine='openpyxl')
        
        final_df.to_excel(excel_buffer, index=False, sheet_name='İstehsalat Hesabatı')
        
        excel_buffer.close()
        output.seek(0)
        
        response = Response(
            output.getvalue(), 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = 'attachment; filename=istehsalat_hesabati.xlsx'
        return response

    except sqlite3.Error as e:
        app.logger.error(f"İstehsalat hesabatı XLSX ixrac edilərkən DB xətası: {e}", exc_info=True)
        flash('Verilənlər bazası xətası səbəbindən hesabat XLSX formatında ixrac edilmədi.', 'danger')
        return redirect(url_for('production_report'))
        
    except Exception as e:
        app.logger.error(f"İstehsalat hesabatı XLSX ixrac edilərkən gözlənilməz xəta: {e}", exc_info=True)
        flash('Gözlənilməz xəta səbəbindən hesabat XLSX formatında ixrac edilmədi.', 'danger')
        return redirect(url_for('production_report'))
        
    finally:
        conn.close()
@app.route('/suppliers', methods=['GET', 'POST'])
@login_required
def suppliers():
    if session['role'] not in ['Admin', 'Operator']:
        flash('Bu səhifəyə giriş icazəniz yoxdur.', 'danger')
        return redirect(url_for('index'))

    conn = get_db_connection()
    try:
        if request.method == 'POST':
            name = request.form.get('name')
            contact_person = request.form.get('contact_person')
            phone = request.form.get('phone')
            email = request.form.get('email')
            address = request.form.get('address')
            tax_id = request.form.get('tax_id')
            bank_account = request.form.get('bank_account')
            notes = request.form.get('notes')
            supplier_id = request.form.get('supplier_id')

            if supplier_id:  
                conn.execute('''UPDATE suppliers SET name=?, contact_person=?, phone=?, email=?, address=?, tax_id=?, bank_account=?, notes=?, updated_at=? WHERE id=?''', (name, contact_person, phone, email, address, tax_id, bank_account, notes, get_current_baku_time(), supplier_id))
                audit_log(conn, session['user_id'], f'Təchizatçı redaktə edildi: {name} (ID:{supplier_id})')
                conn.commit()
                flash('Təchizatçı yeniləndi.', 'success')
                return redirect(url_for('suppliers'))

            else:
                try:
                    conn.execute('''INSERT INTO suppliers (name, contact_person, phone, email, address, tax_id, bank_account, notes, is_active, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (name, contact_person, phone, email, address, tax_id, bank_account, notes, 1, get_current_baku_time()))
                    audit_log(conn, session['user_id'], f'Yeni təchizatçı əlavə edildi: {name}')
                    conn.commit()
                    flash('Təchizatçı əlavə edildi.', 'success')
                    return redirect(url_for('suppliers'))
                except sqlite3.IntegrityError:
                    flash('Bu adla təchizatçı artıq mövcuddur.', 'danger')

        edit_id = request.args.get('edit')
        supplier_edit = None
        if edit_id:
            supplier_edit = conn.execute('SELECT * FROM suppliers WHERE id = ?', (edit_id,)).fetchone()

        suppliers_list = conn.execute('SELECT * FROM suppliers ORDER BY name').fetchall()
        return render_template('suppliers.html', suppliers=suppliers_list, supplier_edit=supplier_edit)
    finally:
        conn.close()


@app.route('/suppliers/toggle/<int:id>')
@login_required
@role_required('Admin')
def toggle_supplier(id):
    conn = get_db_connection()
    try:
        supplier = conn.execute('SELECT * FROM suppliers WHERE id = ?', (id,)).fetchone()
        if not supplier:
            flash('Təchizatçı tapılmadı.', 'danger')
            return redirect(url_for('suppliers'))
        new_status = 0 if supplier['is_active'] else 1
        conn.execute('UPDATE suppliers SET is_active = ? WHERE id = ?', (new_status, id))
        status_text = 'deaktiv edildi' if new_status == 0 else 'aktiv edildi'
        audit_log(conn, session['user_id'], f'Təchizatçı statusu dəyişdirildi: {supplier["name"]} ({status_text})')
        conn.commit()
        flash(f'Təchizatçı uğurla {status_text}.', 'success')
    finally:
        conn.close()
    return redirect(url_for('suppliers'))

    
@app.route('/finance/reports/debts')
@login_required
@role_required('Admin')
def debt_report():
    conn = get_db_connection()
    reports = []
    total_balance = {'receivable': 0, 'debt': 0, 'net': 0}

    try:
   
        debts_query = '''
            SELECT 
                party_name as company_name,
                SUM(CASE WHEN type = 'Receivable' THEN (initial_amount - paid_amount) ELSE 0 END) as total_receivable,
                SUM(CASE WHEN type = 'Debt' THEN (initial_amount - paid_amount) ELSE 0 END) as total_debt
            FROM debts
            WHERE (initial_amount - paid_amount) > 0.01
            GROUP BY party_name
        '''
        companies_from_debts = conn.execute(debts_query).fetchall()
        
        customer_debts_query = '''
            SELECT 
                customer_name as company_name,
                SUM(amount - paid_amount) as total_receivable,
                0 as total_debt
            FROM customer_debts
            WHERE status != 'Ödənilib' AND (amount - paid_amount) > 0.01
            GROUP BY customer_name
        '''
        customer_companies = conn.execute(customer_debts_query).fetchall()
        
        company_map = {}
        
        for company in companies_from_debts + customer_companies:
            name = company['company_name']
            receivable = company['total_receivable'] or 0
            debt = company['total_debt'] or 0
            
            if name not in company_map:
                company_map[name] = {'company_name': name, 'total_receivable': 0, 'total_debt': 0}
            
            company_map[name]['total_receivable'] += receivable
            company_map[name]['total_debt'] += debt
            
        reports = list(company_map.values())
        
        for report in reports:
            report['net_balance'] = report['total_receivable'] - report['total_debt']
            total_balance['receivable'] += report['total_receivable']
            total_balance['debt'] += report['total_debt']
            
        total_balance['net'] = total_balance['receivable'] - total_balance['debt']
        
        reports.sort(key=lambda x: abs(x['net_balance']), reverse=True)


    except sqlite3.Error as e:
        flash(f'Hesabat yüklənərkən xəta: {str(e)}', 'danger')
        app.logger.error(f'Hesabat yüklənərkən SQL xətası: {str(e)}', exc_info=True)
    finally:
        conn.close()

    return render_template('debt_report.html', reports=reports, total_balance=total_balance)
import pandas as pd
from io import BytesIO 
@app.route('/finance/reports/debts/export', methods=['GET'])
@login_required
@role_required('Admin')
def export_debt_report():
    conn = get_db_connection()
    
    try:
        debts_query_consolidated = '''
            SELECT 
                party_name as company_name,
                SUM(CASE WHEN type = 'Receivable' THEN (initial_amount - paid_amount) ELSE 0 END) as total_receivable,
                SUM(CASE WHEN type = 'Debt' THEN (initial_amount - paid_amount) ELSE 0 END) as total_debt
            FROM debts
            WHERE (initial_amount - paid_amount) > 0.01
            GROUP BY party_name
        '''
        companies_from_debts = pd.read_sql_query(debts_query_consolidated, conn)
        
        customer_debts_query_consolidated = '''
            SELECT 
                customer_name as company_name,
                SUM(amount - paid_amount) as total_receivable,
                0 as total_debt
            FROM customer_debts
            WHERE status != 'Ödənilib' AND (amount - paid_amount) > 0.01
            GROUP BY customer_name
        '''
        customer_companies = pd.read_sql_query(customer_debts_query_consolidated, conn)
        
        df_all = pd.concat([companies_from_debts, customer_companies], ignore_index=True)
        
        df_consolidated = df_all.groupby('company_name', as_index=False).agg(
            total_receivable=('total_receivable', 'sum'),
            total_debt=('total_debt', 'sum')
        )
        
        df_consolidated['Net Balans (Alacaq - Borc)'] = df_consolidated['total_receivable'] - df_consolidated['total_debt']
        df_consolidated.columns = [
            'Tərəfin Adı',
            'Ümumi Alacaq (Bizə Borcludur)',
            'Ümumi Borc (Biz Borcluyuq)',
            'Xalis Balans (Alacaq - Borc)'
        ]
        
        debts_query_detailed = '''
            SELECT
                id,
                party_name AS Tərəf,
                type AS Növ,
                initial_amount AS İlkin_Məbləğ,
                paid_amount AS Ödənilmiş_Məbləğ,
                (initial_amount - paid_amount) AS Qalıq_Məbləğ,
                'Əsas Borc/Alacaq' AS Mənbə
            FROM debts
            WHERE (initial_amount - paid_amount) > 0.01
            ORDER BY party_name
        '''
        df_debts_detailed = pd.read_sql_query(debts_query_detailed, conn)
        
        customer_debts_query_detailed = '''
            SELECT
                id,
                customer_name AS Tərəf,
                'Alacaq' AS Növ,
                amount AS İlkin_Məbləğ,
                paid_amount AS Ödənilmiş_Məbləğ,
                (amount - paid_amount) AS Qalıq_Məbləğ,
                'Müştəri Borcu' AS Mənbə
            FROM customer_debts
            WHERE status != 'Ödənilib' AND (amount - paid_amount) > 0.01
            ORDER BY customer_name
        '''
        df_customer_debts_detailed = pd.read_sql_query(customer_debts_query_detailed, conn)
        
        df_detailed_all = pd.concat([df_debts_detailed, df_customer_debts_detailed], ignore_index=True)

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        df_consolidated.to_excel(writer, sheet_name='Şirkətlər üzrə ümumi hesabat', index=False)
        
        df_detailed_all.to_excel(writer, sheet_name='Ətraflı Məlumatlar', index=False)
        
        writer.close()
        
        output.seek(0)
        
        current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'Borc_Teleb_Hesabati_{current_time}.xlsx'
        
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except ImportError:
        flash('Pandas və ya Openpyxl kitabxanaları yüklənməyib. Zəhmət olmasa "pip install pandas openpyxl" əmrini icra edin.', 'danger')
        return redirect(url_for('debt_report'))
    except Exception as e:
        flash(f'Hesabatın eksportu zamanı xəta: {str(e)}', 'danger')
        app.logger.error(f'Hesabat eksport xətası: {str(e)}', exc_info=True)
        return redirect(url_for('debt_report'))
    finally:
        conn.close()

@app.cli.command('init-db')
def init_db_command():
    init_db()
    print('Database initialized.')

@app.route('/transaction_history')
@login_required
def transaction_history():
    if session.get('role') not in ['Admin', 'Operator']:
        flash('Bu səhifəyə giriş icazəniz yoxdur.', 'danger')
        return redirect(url_for('index'))
    
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    transaction_type = request.args.get('type', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    employee_id = request.args.get('employee_id', '')
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    product_id = request.args.get('product_id', '')
    
    per_page = 50  # Number of transactions per page
    
    try:
        with DatabaseConnection() as conn:
            # Get employees and products for dropdowns
            employees = conn.execute('SELECT id, username FROM employees ORDER BY username').fetchall()
            products = conn.execute('SELECT id, name FROM products WHERE is_active = 1 ORDER BY name').fetchall()
            
            # Base query for counting and fetching transactions
            base_query = '''
                FROM transactions t
                LEFT JOIN products p ON t.product_id = p.id
                LEFT JOIN employees e ON t.employee_id = e.id
            '''
            
            # Initialize conditions and parameters
            where_conditions = []
            params = []
            
            # Search term condition
            if search_query:
                search_term = f'%{search_query}%'
                search_conditions = [
                    'LOWER(p.name) LIKE LOWER(?)',
                    'LOWER(e.username) LIKE LOWER(?)',
                    'LOWER(t.customer_name) LIKE LOWER(?)',
                    'LOWER(t.note) LIKE LOWER(?)',
                    'LOWER(t.type) LIKE LOWER(?)'
                ]
                where_conditions.append('(' + ' OR '.join(search_conditions) + ')')
                params.extend([search_term] * 5)
            
            # Transaction type filter
            if transaction_type in ['purchase', 'sale']:
                where_conditions.append('t.type = ?')
                params.append(transaction_type)
            
            # Date range filter
            if start_date:
                where_conditions.append('DATE(t.timestamp) >= ?')
                params.append(start_date)
            if end_date:
                where_conditions.append('DATE(t.timestamp) <= ?')
                params.append(end_date)
            
            # Employee filter
            if employee_id and employee_id.isdigit():
                where_conditions.append('t.employee_id = ?')
                params.append(int(employee_id))
            
            # Amount range filter
            if min_amount:
                try:
                    where_conditions.append('t.total >= ?')
                    params.append(float(min_amount))
                except ValueError:
                    pass
            if max_amount:
                try:
                    where_conditions.append('t.total <= ?')
                    params.append(float(max_amount))
                except ValueError:
                    pass
            
            # Product filter
            if product_id and product_id.isdigit():
                where_conditions.append('t.product_id = ?')
                params.append(int(product_id))
            
            # Build the WHERE clause
            where_clause = ' WHERE ' + ' AND '.join(where_conditions) if where_conditions else ''
            
            # Debug info
            app.logger.debug(f'Search conditions: {where_conditions}')
            app.logger.debug(f'Search params: {params}')
            
            # Get total number of transactions matching search criteria
            count_query = 'SELECT COUNT(*) ' + base_query + where_clause
            total_transactions = conn.execute(count_query, params).fetchone()[0]
            total_pages = max(1, (total_transactions + per_page - 1) // per_page)
            
            # Ensure page is within valid range
            page = max(1, min(page, total_pages))
            
            # Calculate offset for pagination
            offset = (page - 1) * per_page
            
            # Get transactions for the current page with search and pagination
            query = f'''
                SELECT t.*, p.name as product_name, p.unit as product_unit,
                       e.username as employee_name
                {base_query}
                {where_clause}
                ORDER BY t.timestamp DESC
                LIMIT ? OFFSET ?
            '''
            
            # Execute the query with all parameters
            app.logger.debug(f'Executing query: {query}')
            app.logger.debug(f'With params: {params + [per_page, offset]}')
            
            transactions_data = conn.execute(query, params + [per_page, offset]).fetchall()
            app.logger.debug(f'Found {len(transactions_data)} transactions')
            
            # Convert Row objects to dictionaries and ensure datetime objects
            transactions = []
            for row in transactions_data:
                row_dict = dict(row)
                # Convert timestamp string to datetime object if it's a string
                if isinstance(row_dict.get('timestamp'), str):
                    try:
                        row_dict['timestamp'] = datetime.datetime.strptime(
                            row_dict['timestamp'], '%Y-%m-%d %H:%M:%S'
                        )
                    except (ValueError, TypeError):
                        # If conversion fails, keep the original value
                        pass
                transactions.append(row_dict)
            
            return render_template('transaction_history.html',
                               transactions=transactions,
                               current_page=page,
                               total_pages=total_pages,
                               total_transactions=total_transactions,
                               search_query=search_query,
                               employees=employees,
                               products=products)
    except Exception as e:
        app.logger.error(f'Transaction history error: {e}', exc_info=True)
        flash('Tarixçə yüklənərkən xəta baş verdi.', 'danger')
        return redirect(url_for('index'))

with app.app_context():
    init_db()

if __name__ == '__main__':
    # Log faylının yaradılması
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('app_debug.log'),
            logging.StreamHandler()
        ]
    )
    app.logger.setLevel(logging.DEBUG)
    app.run(host='0.0.0.0', port=5000, debug=True)